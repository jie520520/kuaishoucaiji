# -*- coding: utf-8 -*-
"""
颜阿娇 - 批量提取达人联系方式 v3
=================================
从采集结果Excel中读取达人列表，逐个打开详情页，
点击"查看联系方式"，提取手机号和微信号，输出到新Excel。

技术路线（用户确认）：
  - 点击"查看联系方式"后，手机号和微信号都可以直接文本复制
  - 无需OCR，纯DOM文本提取即可

使用方式：
  python extract_contacts.py --input 采集结果/颜阿娇_达人_健康_20260713_124840.xlsx
  python extract_contacts.py --input 采集结果/颜阿娇_达人_健康_20260713_124840.xlsx --max 20
  python extract_contacts.py --id 535256404  (单个测试)
"""

import sys
import io
import time
import json
import re
import argparse
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    try:
        if hasattr(sys.stdout, 'buffer') and not isinstance(sys.stdout, io.TextIOWrapper):
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        if hasattr(sys.stderr, 'buffer') and not isinstance(sys.stderr, io.TextIOWrapper):
            sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(line_buffering=True)
    except Exception:
        pass  # 作为模块导入或管道模式时忽略

_START_TIME = time.time()


def _ts() -> str:
    return datetime.now().strftime("%H:%M:%S")


def log(msg: str, end: str = "\n"):
    text = f"[{_ts()}] {msg}"
    print(text, end=end, flush=True)


from playwright.sync_api import sync_playwright
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ============================================================
# 配置
# ============================================================

class Config:
    BASE_DIR = Path(__file__).parent
    USER_DATA = BASE_DIR / ".kuaishou_browser_data"
    OUTPUT_DIR = BASE_DIR / "采集结果"
    DAREN_SQUARE_URL = "https://cps.kwaixiaodian.com/zone/daren-match/daren-square-pro"
    DETAIL_URL_FMT = "https://cps.kwaixiaodian.com/zone/daren-match/daren-detail?promoterId={pid}"

    # 提取限制
    PAGE_DELAY = 5          # 每人间隔（秒）
    MAX_PER_SESSION = 500   # 每次最多提取人数（官方每日上限 500）
    BATCH_SIZE = 10         # 每批人数（中间暂停一下）

    # 并发/频率限制：保守策略
    LONG_PAUSE_EVERY = 20   # 每N人后额外暂停
    LONG_PAUSE_DURATION = 15  # 额外暂停时长（秒）

    # 自动暂停恢复：连续N个达人无联系方式 → 判断触达每日上限 → 等30分钟恢复
    AUTO_PAUSE_CONSECUTIVE_EMPTY = 5   # 连续N个达人无联系方式触发暂停（5=较合理，避免偶发空结果误触发）
    AUTO_PAUSE_MINUTES = 30            # 暂停分钟后自动恢复


# ============================================================
# 单个达人联系方式提取
# ============================================================

def extract_one(page, promoter_id: str, nickname: str = "") -> dict:
    """
    从详情页提取单个达人的手机号和微信号。
    返回 {"promoterId", "nickname", "phone", "wechat", "error"}
    """
    result = {
        "promoterId": promoter_id,
        "nickname": nickname,
        "phone": "",
        "wechat": "",
        "error": "",
    }

    detail_url = Config.DETAIL_URL_FMT.format(pid=promoter_id)

    try:
        # 1. 导航到详情页
        page.goto(detail_url, timeout=30000, wait_until="domcontentloaded")
        time.sleep(4)

        # 检查页面是否有效
        body_text = page.locator("body").inner_text(timeout=3000)
        if "未找到" in body_text or "404" in body_text:
            result["error"] = "详情页不存在"
            return result

        # 2. 点击"查看联系方式"
        contact_btn = page.locator('text=查看联系方式').first
        if contact_btn.count() > 0:
            try:
                contact_btn.click(timeout=5000)
                time.sleep(3)
            except Exception as e:
                result["error"] = f"点击查看联系方式失败: {e}"
                return result
        else:
            # 可能已经展开了，或者没有联系方式
            if "手机" in body_text and "微信" in body_text:
                pass  # 已经显示
            else:
                result["error"] = "无查看联系方式按钮"
                return result

        # 3. 获取最新页面文本
        time.sleep(1)
        body_text = page.locator("body").inner_text(timeout=3000)

        # 4. 提取手机号（优先匹配"手机号+86xxx"格式，再兜底全局正则）
        phones = []
        # 优先：匹配"手机号"后面的+86格式
        phone_86 = re.findall(r'手机号\+?86(\d{11})', body_text)
        if phone_86:
            phones = phone_86
        else:
            # 兜底：全局11位手机号
            phones = re.findall(r'1[3-9]\d{9}', body_text)
        # 过滤：排除含4个以上连续相同数字（排除系统ID截断如14000011687）
        phones = [p for p in set(phones) if not re.search(r'(\d)\1{3,}', p)]
        # 只取真正11位的（排除更长ID的截断）
        phones = [p for p in phones if len(p) == 11]
        result["phone"] = phones[0] if phones else ""

        # 5. 提取微信号（多种格式）
        # 快手页面常见格式："微信号ZJ20258686"（无分隔符）、"微信号：xxx"、"微信号 xxx"
        wechats = []
        for pat in [
            r'微信号\s*[：:]\s*([a-zA-Z0-9_-]{4,40})',
            r'微信\s*[：:]\s*([a-zA-Z0-9_-]{4,40})',
            r'微信号\s+([a-zA-Z0-9_-]{4,40})',
            r'微信\s+([a-zA-Z0-9_-]{4,40})',
            # 无分隔符兜底（"微信号xxx"直接连在一起）
            r'微信号([a-zA-Z][a-zA-Z0-9_-]{3,39})',
            r'微信([a-zA-Z][a-zA-Z0-9_-]{3,39})',
        ]:
            found = re.findall(pat, body_text)
            wechats.extend(found)
        # 过滤常见的误匹配
        exclude = {"该用户", "暂无", "null", "undefined", "wxid", "gh_"}
        wechats = list(set(w for w in wechats if w not in exclude))
        result["wechat"] = wechats[0] if wechats else ""

        if not result["phone"] and not result["wechat"]:
            result["error"] = "未提取到手机号或微信号"

    except Exception as e:
        result["error"] = str(e)[:100]

    return result


# ============================================================
# 浏览器生命周期 & 批量提取
# ============================================================

def _open_browser():
    """打开浏览器并返回 (p, ctx, page)，同时导航空到达人广场确认登录。"""
    log("🚀 启动浏览器（使用已有登录态）...")
    p = sync_playwright().start()
    ctx = p.chromium.launch_persistent_context(
        user_data_dir=str(Config.USER_DATA),
        headless=False,
        args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        viewport={"width": 1440, "height": 900},
        locale="zh-CN",
    )
    page = ctx.new_page()
    page.goto(Config.DAREN_SQUARE_URL, timeout=60000, wait_until="domcontentloaded")
    time.sleep(5)

    if "login" in page.url.lower():
        log("⚠️  需要扫码登录快手，请在浏览器中扫码...")
        for i in range(120):
            time.sleep(2)
            if "login" not in page.url.lower():
                log("🔓 登录成功！")
                time.sleep(3)
                break
        else:
            log("❌ 登录超时")
            ctx.close()
            p.stop()
            return None, None, None

    return p, ctx, page


def _close_browser(p, ctx):
    """安全关闭浏览器"""
    try:
        ctx.close()
    except Exception:
        pass
    try:
        p.stop()
    except Exception:
        pass


def batch_extract(input_excel: str, max_count: int = None,
                  start_from: int = 0, callback=None):
    """
    批量从Excel中读取达人列表，逐个提取联系方式。

    特性：
    - 智能跳过已有联系方式的达人（断点续提）
    - 连续N个达人无联系方式 → 自动暂停30分钟 → 恢复（模拟人工等待触达上限恢复）

    参数：
        input_excel: 采集结果Excel路径
        max_count: 最多提取人数（None=不限制，默认500上限）
        start_from: 从第几个待提取达人开始（用于断点续传）
        callback: 回调函数 callback(row_index, result, progress_info)

    返回：
        list of dict: 所有提取结果
    """
    log(f"📂 读取输入文件: {input_excel}")
    wb = load_workbook(input_excel, read_only=True)
    ws = wb.active

    # 读取表头，找到关键列
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header = [str(h).strip() if h else "" for h in header]

    pid_col = None
    name_col = None
    phone_col = None
    wechat_col = None
    for i, h in enumerate(header):
        hl = h.lower()
        if "快手id" in hl or "promoter" in hl or h == "快手ID":
            pid_col = i
        if "达人名称" in hl or "昵称" in hl or h == "达人名称":
            name_col = i
        if "手机号" in h:
            phone_col = i
        if "微信号" in h:
            wechat_col = i

    if pid_col is None:
        log("❌ 未找到'快手ID'或'promoterId'列")
        log(f"   表头: {header}")
        wb.close()
        return []

    log(f"   promoterId列: {pid_col} ({header[pid_col]})")
    if name_col is not None:
        log(f"   昵称列: {name_col} ({header[name_col]})")

    # 读取全部数据行，标记哪些已有联系方式
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    if not all_rows:
        log("❌ Excel中没有数据行")
        return []

    # 过滤：跳过已有手机号或微信号的行
    rows_to_extract = []  # (excel行号1-based, 行数据)
    already_has = 0
    for idx, row in enumerate(all_rows):
        has_contact = False
        if phone_col is not None and row[phone_col]:
            val = str(row[phone_col]).strip()
            if val and val not in ("None", ""):
                has_contact = True
        if wechat_col is not None and row[wechat_col]:
            val = str(row[wechat_col]).strip()
            if val and val not in ("None", ""):
                has_contact = True
        if has_contact:
            already_has += 1
        else:
            rows_to_extract.append((idx + 2, row))  # row index in Excel (1-based)

    total = len(all_rows)
    need_extract = len(rows_to_extract)

    if already_has > 0:
        log(f"   ⏭ 已有联系方式: {already_has} 人 (自动跳过)")
    log(f"   🎯 待提取: {need_extract} 人")

    if need_extract == 0:
        log("✅ 所有达人联系方式已齐全，无需提取")
        return []

    if max_count is None:
        max_count = min(Config.MAX_PER_SESSION, need_extract)
    else:
        max_count = min(max_count, need_extract)

    if start_from >= need_extract:
        log(f"❌ start_from={start_from} 超出待提取范围")
        return []

    effective_total = min(max_count, need_extract - start_from)
    log(f"📊 本次提取: {start_from+1} ~ {start_from+effective_total} / {need_extract} (共{effective_total}人)")

    # 启动浏览器
    p, ctx, page = _open_browser()
    if page is None:
        return []

    results = []
    success_count = 0
    consecutive_empty = 0          # 连续无联系方式的计数
    pause_count = 0                # 暂停次数

    # 使用 while 循环支持暂停恢复
    idx = start_from
    while idx < start_from + effective_total:
        row_excel_num, row = rows_to_extract[idx]
        promoter_id = str(row[pid_col]) if row[pid_col] else ""
        nickname = str(row[name_col]) if name_col is not None and len(row) > name_col and row[name_col] else ""

        if not promoter_id or promoter_id.upper() == "NONE":
            log(f"  ⏭ [{idx+1}/{need_extract}] 跳过：无ID")
            results.append({"promoterId": "", "nickname": nickname, "phone": "", "wechat": "", "error": "无ID"})
            idx += 1
            continue

        # 进度显示
        done_so_far = idx - start_from + 1
        remaining = effective_total - done_so_far
        progress = f"[{idx+1}/{need_extract}] 已提取{success_count}人, 剩余{remaining}人"

        log(f"\n{'='*50}")
        log(f"  🎯 {progress}")
        log(f"  📋 {nickname} (ID:{promoter_id})")

        # 提取
        result = extract_one(page, promoter_id, nickname)

        if result["phone"] or result["wechat"]:
            success_count += 1
            consecutive_empty = 0
            emoji = "✅"
        else:
            consecutive_empty += 1
            emoji = "⚠️ "

        log(f"  {emoji} 手机: {result['phone'] or '---'}  微信: {result['wechat'] or '---'}")
        if result["error"]:
            log(f"     错误: {result['error']}")

        results.append(result)

        # 回调
        if callback:
            callback(idx, result, progress)

        # 保存进度：JSON 每步存，Excel 每5步批量写（防卡锁）
        _save_progress(results, input_excel)
        if done_so_far % 5 == 0 or done_so_far == effective_total or result["phone"] or result["wechat"]:
            try:
                _save_final_excel(results, input_excel)
            except Exception as e:
                log(f"  ⚠️ Excel保存失败（JSON已存）: {e}")

        idx += 1  # 指针前移（无论失败与否，不重试同一个达人）

        # ========================================
        # 检测：连续N个无联系方式 → 自动暂停恢复
        # ========================================
        if consecutive_empty >= Config.AUTO_PAUSE_CONSECUTIVE_EMPTY:
            pause_count += 1
            log(f"\n{'⚠️'*20}")
            log(f"⚠️  连续 {consecutive_empty} 个达人无联系方式！")
            log(f"⚠️  疑似触达快手每日查询上限，触发自动暂停...")
            log(f"{'⚠️'*20}")

            # 先保存已有结果到 Excel
            _save_final_excel(results, input_excel)

            # 不关浏览器！保持登录会话存活，避免恢复时需要人工扫码
            page.goto(Config.DAREN_SQUARE_URL, timeout=30000, wait_until="domcontentloaded")
            time.sleep(2)

            wait_seconds = Config.AUTO_PAUSE_MINUTES * 60
            wait_end = datetime.now().timestamp() + wait_seconds
            log(f"⏰ 暂停 {Config.AUTO_PAUSE_MINUTES} 分钟（浏览器保持打开，无需重新登录）")
            log(f"   预计 {datetime.fromtimestamp(wait_end).strftime('%H:%M:%S')} 恢复...")
            log(f"   (已提取 {len(results)} 人，成功 {success_count} 人，第 {pause_count} 次暂停)")

            # 分段等待，每30秒输出一次剩余时间
            while wait_seconds > 0:
                chunk = min(30, wait_seconds)
                time.sleep(chunk)
                wait_seconds -= chunk
                if wait_seconds > 0:
                    log(f"   ⏳ 剩余约 {wait_seconds // 60} 分 {wait_seconds % 60} 秒...")

            # 快速验证登录态（导航到广场确认）
            log(f"\n{'='*50}")
            log(f"🔄 第 {pause_count} 次恢复：验证登录态...")
            page.goto(Config.DAREN_SQUARE_URL, timeout=30000, wait_until="domcontentloaded")
            time.sleep(3)

            if "login" in page.url.lower():
                log("⚠️  登录态已过期，需要重新扫码（等待 30 秒自动检测）...")
                for _ in range(15):
                    time.sleep(2)
                    if "login" not in page.url.lower():
                        log("🔓 登录已恢复！")
                        break
                else:
                    log("❌ 登录态已过期且无法自动恢复，请手动扫码后继续")
                    log("   浏览器窗口已打开，扫码后程序将继续...")
                    for _ in range(60):
                        time.sleep(2)
                        if "login" not in page.url.lower():
                            log("🔓 登录成功，继续提取！")
                            break
                    else:
                        log("❌ 登录超时，停止提取")
                        break

            consecutive_empty = 0  # 重置计数器
            log(f"✅ 已恢复，从第 {idx-start_from+1}/{effective_total} 个达人继续")
            continue  # 不需要 sleep，直接下一个

        # 频率控制（正常模式）
        if idx < start_from + effective_total:
            delay = Config.PAGE_DELAY
            log(f"  ⏳ 等待 {delay} 秒...")
            time.sleep(delay)

            # 每N人额外暂停
            if done_so_far % Config.LONG_PAUSE_EVERY == 0:
                log(f"  🫁 每{Config.LONG_PAUSE_EVERY}人额外暂停 {Config.LONG_PAUSE_DURATION} 秒...")
                time.sleep(Config.LONG_PAUSE_DURATION)

    # 关闭浏览器
    _close_browser(p, ctx)

    # 保存最终结果
    output_file = _save_final_excel(results, input_excel)
    log(f"\n{'='*50}")
    log(f"🎉 提取完成！")
    log(f"   成功: {success_count}/{len(results)}")
    if pause_count > 0:
        log(f"   自动暂停恢复: {pause_count} 次")
    log(f"   输出: {output_file}")

    return results


# ============================================================
# 进度保存 & 最终输出
# ============================================================

def _save_progress(results: list, input_excel: str):
    """增量保存为JSON（用于断点续传恢复）"""
    input_stem = Path(input_excel).stem
    progress_file = Config.OUTPUT_DIR / f"{input_stem}_联系方式_progress.json"
    with open(progress_file, "w", encoding="utf-8") as f:
        json.dump({
            "input": str(input_excel),
            "extracted": len(results),
            "timestamp": datetime.now().isoformat(),
            "results": results,
        }, f, ensure_ascii=False, indent=2)


def _save_final_excel(results: list, input_excel: str):
    """将提取的手机号和微信号回写到原始Excel（无论有无联系方式都标记）"""
    input_path = Path(input_excel)

    # 构建结果查找表：{promoterId: (phone, wechat)}
    # 有联系方式的：写实际号码；无联系方式的：phone写"无"作标记，避免下次重复提取
    lookup = {}
    for r in results:
        pid = str(r.get("promoterId", "")).strip()
        if not pid:
            continue
        phone = r.get("phone") or ""
        wechat = r.get("wechat") or ""
        # 无联系方式的用"无"标记，过滤逻辑会识别并跳过
        if not phone and not wechat:
            phone = "无"
        lookup[pid] = (phone, wechat)

    if not lookup:
        log("⚠️ 没有可回写的数据")
        return ""

    # 打开原始文件（读写模式）
    wb = load_workbook(input_excel)
    ws = wb.active

    # 读取表头，找关键列
    pid_col = None
    phone_col = None
    wechat_col = None
    for col_idx, cell in enumerate(ws[1], 1):
        h = str(cell.value).strip() if cell.value else ""
        if "快手ID" in h or "promoter" in h.lower():
            pid_col = col_idx
        if h == "手机号":
            phone_col = col_idx
        if h == "微信号":
            wechat_col = col_idx

    if pid_col is None:
        log("❌ 未找到快手ID列，无法回写到原文件")
        wb.close()
        return ""

    # 如果列不存在则新增，存在则复用已有列号
    need_new_phone = (phone_col is None)
    need_new_wechat = (wechat_col is None)
    max_col = ws.max_column
    if need_new_phone and need_new_wechat:
        phone_col = max_col + 1
        wechat_col = max_col + 2
    elif need_new_phone:
        phone_col = max_col + 1
    elif need_new_wechat:
        wechat_col = max_col + 1

    # 样式
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="微软雅黑", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 写新表头（仅新建列需要）
    for col, title, is_new in [(phone_col, "手机号", need_new_phone), (wechat_col, "微信号", need_new_wechat)]:
        if is_new:
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    # 逐行回填（仅更新本次新提取的，已有数据的保留不动）
    filled = 0
    skipped = 0
    max_row = ws.max_row
    for row_idx in range(2, max_row + 1):
        pid_cell = ws.cell(row=row_idx, column=pid_col)
        pid = str(pid_cell.value).strip() if pid_cell.value else ""

        # 检查该行是否已有联系方式（已有则跳过，保留原数据）
        existing_phone = ws.cell(row=row_idx, column=phone_col).value
        existing_wechat = ws.cell(row=row_idx, column=wechat_col).value
        has_existing = bool(
            (existing_phone and str(existing_phone).strip() not in ("", "None"))
            or (existing_wechat and str(existing_wechat).strip() not in ("", "None"))
        )
        if has_existing:
            skipped += 1
            continue

        # 从本次提取结果查找
        phone_val, wechat_val = lookup.get(pid, ("", ""))
        has_real_contact = bool((phone_val and phone_val != "无") or wechat_val)
        if has_real_contact:
            filled += 1

        for col, val in [(phone_col, phone_val), (wechat_col, wechat_val)]:
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = center_align

        # 颜色：绿=有联系方式，红=无/未提取到
        row_fill = green_fill if has_real_contact else red_fill
        ws.cell(row=row_idx, column=phone_col).fill = row_fill
        ws.cell(row=row_idx, column=wechat_col).fill = row_fill

    # 列宽
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(phone_col)].width = 16
    ws.column_dimensions[get_column_letter(wechat_col)].width = 22

    # 保存回原文件
    wb.save(input_excel)
    wb.close()

    log(f"💾 已回写到原文件: {input_path.name}")
    log(f"   本次新增: {filled} 人")
    if skipped > 0:
        log(f"   已有跳过: {skipped} 人 (保留原数据)")
    return str(input_path)


# ============================================================
# 单个结果精准追加（用于实时保存，O(1)行写入）
# ============================================================

# 缓存：{input_excel: (pid_col, phone_col, wechat_col)}，避免每次检测列
_col_cache = {}


def _save_one_to_excel(result: dict, input_excel: str):
    """将单个提取结果精准写入Excel对应行。
    
    与 _save_final_excel 的区别：
    - _save_final_excel 遍历全部行再写 → O(n²) I/O，日志混淆
    - _save_one_to_excel 按PID精准定位→只写一行 → O(1)，日志清晰
    """
    pid = str(result.get("promoterId", "")).strip()
    if not pid:
        return ""

    phone = result.get("phone") or ""
    wechat = result.get("wechat") or ""
    if not phone and not wechat:
        phone = "无"

    has_real = bool((phone and phone != "无") or wechat)

    input_path = Path(input_excel)
    key = str(input_path.resolve())
    wb = load_workbook(input_excel)
    ws = wb.active

    # 列检测（优先用缓存）
    cached = _col_cache.get(key)
    if cached:
        pid_col, phone_col, wechat_col = cached
    else:
        pid_col = phone_col = wechat_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            h = str(cell.value).strip() if cell.value else ""
            if "快手ID" in h or "promoter" in h.lower():
                pid_col = col_idx
            if h == "手机号":
                phone_col = col_idx
            if h == "微信号":
                wechat_col = col_idx
        if pid_col is None:
            wb.close()
            return ""

        need_new_phone = (phone_col is None)
        need_new_wechat = (wechat_col is None)
        max_col = ws.max_column
        if need_new_phone and need_new_wechat:
            phone_col = max_col + 1
            wechat_col = max_col + 2
        elif need_new_phone:
            phone_col = max_col + 1
        elif need_new_wechat:
            wechat_col = max_col + 1

        # 写新表头
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        for col, title, is_new in [(phone_col, "手机号", need_new_phone), (wechat_col, "微信号", need_new_wechat)]:
            if is_new:
                cell = ws.cell(row=1, column=col, value=title)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        _col_cache[key] = (pid_col, phone_col, wechat_col)

    # 样式
    body_font = Font(name="微软雅黑", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 按PID查找目标行
    found_row = None
    max_row = ws.max_row
    for row_idx in range(2, max_row + 1):
        cell_val = str(ws.cell(row=row_idx, column=pid_col).value or "").strip()
        if cell_val == pid:
            found_row = row_idx
            break

    if found_row is None:
        wb.close()
        return ""

    # 已有联系方式的不覆盖
    existing_phone = ws.cell(row=found_row, column=phone_col).value
    existing_wechat = ws.cell(row=found_row, column=wechat_col).value
    has_existing = bool(
        (existing_phone and str(existing_phone).strip() not in ("", "None"))
        or (existing_wechat and str(existing_wechat).strip() not in ("", "None"))
    )
    if has_existing:
        wb.close()
        return ""

    row_fill = green_fill if has_real else red_fill

    # 写入手机号
    ph_cell = ws.cell(row=found_row, column=phone_col)
    ph_cell.value = phone
    ph_cell.font = body_font
    ph_cell.border = thin_border
    ph_cell.alignment = center_align
    ph_cell.fill = row_fill

    # 写入微信号
    wx_cell = ws.cell(row=found_row, column=wechat_col)
    wx_cell.value = wechat
    wx_cell.font = body_font
    wx_cell.border = thin_border
    wx_cell.alignment = center_align
    wx_cell.fill = row_fill

    # 列宽
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(phone_col)].width = 16
    ws.column_dimensions[get_column_letter(wechat_col)].width = 22

    wb.save(input_excel)
    wb.close()

    tag = "📱" if has_real else "⛔"
    log(f"{tag} [{pid}] → {'phone=' + phone if has_real else '无（已标记）'}")
    return str(input_path)


# ============================================================
# 命令行入口
# ============================================================

def find_latest_excel() -> str:
    """在采集结果目录找最新的Excel"""
    excels = list(Config.OUTPUT_DIR.glob("颜阿娇_达人_*_*.xlsx"))
    # 排除已含"联系方式"的
    excels = [f for f in excels if "联系方式" not in f.stem]
    if not excels:
        return ""
    excels.sort(key=lambda f: f.stat().st_mtime, reverse=True)
    return str(excels[0])


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="颜阿娇 - 批量提取达人联系方式")
    parser.add_argument("--input", type=str, default="",
                        help="输入Excel路径（不指定则自动选最新）")
    parser.add_argument("--max", type=int, default=None,
                        help="最多提取人数（默认50）")
    parser.add_argument("--id", type=str, default="",
                        help="测试单个达人ID")
    parser.add_argument("--delay", type=int, default=Config.PAGE_DELAY,
                        help=f"每人间隔秒数（默认{Config.PAGE_DELAY}）")
    args = parser.parse_args()

    if args.id:
        # 单个测试
        log(f"🔍 测试模式：提取达人 {args.id}")
        p = sync_playwright().start()
        ctx = p.chromium.launch_persistent_context(
            user_data_dir=str(Config.USER_DATA),
            headless=False,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
            viewport={"width": 1440, "height": 900},
            locale="zh-CN",
        )
        page = ctx.new_page()
        page.goto(Config.DAREN_SQUARE_URL, timeout=60000, wait_until="domcontentloaded")
        time.sleep(5)

        if "login" in page.url.lower():
            log("请在浏览器中扫码登录...")
            for i in range(120):
                time.sleep(2)
                if "login" not in page.url.lower():
                    log("登录成功！")
                    time.sleep(3)
                    break
            else:
                log("登录超时")
                ctx.close(); p.stop()
                sys.exit(1)

        result = extract_one(page, args.id)
        print(f"\n结果: {json.dumps(result, ensure_ascii=False, indent=2)}")
        ctx.close()
        p.stop()
        sys.exit(0)

    # 批量模式
    Config.PAGE_DELAY = args.delay

    input_file = args.input or find_latest_excel()
    if not input_file:
        log("❌ 未找到输入Excel，请用 --input 指定")
        sys.exit(1)

    if not Path(input_file).exists():
        log(f"❌ 文件不存在: {input_file}")
        sys.exit(1)

    batch_extract(input_file, max_count=args.max)
