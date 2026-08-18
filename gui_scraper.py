# -*- coding: utf-8 -*-
"""
颜阿娇快手达人采集工具 - GUI 版 v11.5
====================================
基于 yanajiao_scraper.py 的图形界面版本。

v11.5: 筛选条件改为"平台先筛后拉"
  - 采集前在达人广场页面真实点击筛选条件（达人性别/地域/粉丝年龄/粉性/城市等级）
  - 快手平台直接返回筛选后的达人，不再"拉全部→本地筛"的过度查询
  - 拦截页面请求参数翻页，保证与页面筛选完全一致

功能:
  - 可视化选择采集标签 (37个, 推荐标签高亮)
  - 多选筛选条件：达人性别/地域/粉丝年龄/粉性/城市等级
  - 实时进度条 + 彩色日志
  - 可调节采集速度 (页间延迟/标签间延迟/每页数量)
  - 一键启动/停止, 打开输出文件夹
  - 极速模式 (0.5秒/页, 50条/页)

使用:
  双击 启动GUI.bat
  或 python gui_scraper.py
"""

import os
import sys
import io
import time
import json
import queue
import threading
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    try:
        # 必须重配编码为 utf-8，否则 emoji 日志（🛡⚠️💾）在 GBK 控制台/日志文件下
        # 会触发 UnicodeEncodeError，导致后台线程崩溃（表现为 GUI 卡在"启动浏览器中..."）
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8', errors='replace', line_buffering=True)
        if hasattr(sys.stderr, 'reconfigure'):
            sys.stderr.reconfigure(encoding='utf-8', errors='replace', line_buffering=True)
    except Exception:
        pass

import traceback
import ctypes


def _show_err(title, text):
    """不依赖 tkinter 的原生 Windows 错误弹窗（ctypes），保证任何崩溃都可见、可截图"""
    try:
        ctypes.windll.user32.MessageBoxW(0, str(text), str(title), 0x10)
    except Exception:
        pass


try:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
except Exception as _e:
    # tkinter 自身都加载不了（最常见：python 缺 tcl/tk 运行时）
    _msg = "无法加载 tkinter（GUI 库）：%r\n" % (_e,)
    sys.stderr.write(_msg)
    _crash = Path(__file__).parent.joinpath("gui_crash.log")
    try:
        _crash.write_text(
            "时间: %s\n\n%s" % (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), _msg),
            encoding="utf-8",
        )
    except Exception:
        pass
    _show_err("启动失败 - tkinter", _msg)  # 原生弹窗，必定可见
    try:
        os.startfile(str(_crash))
    except Exception:
        pass
    sys.exit(1)

# 导入采集器模块 (同目录)
sys.path.insert(0, str(Path(__file__).parent))
try:
    import yanajiao_scraper as ys
    import extract_contacts as ec
except Exception:
    # 采集模块导入失败（依赖缺失 / 模块内语法错误等）：弹窗 + 写日志，不让它静默闪退
    _tb = traceback.format_exc()
    try:
        _rk = tk.Tk()
        _rk.withdraw()
        messagebox.showerror("启动失败", "导入采集模块失败：\n\n%s" % _tb)
        _rk.destroy()
    except Exception:
        sys.stderr.write(_tb + "\n")
    _crash = Path(__file__).parent.joinpath("gui_crash.log")
    try:
        _crash.write_text(
            "时间: %s\n\n%s" % (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), _tb),
            encoding="utf-8",
        )
    except Exception:
        pass
    _show_err("启动失败 - 导入模块", _tb)  # 原生弹窗，必定可见
    try:
        os.startfile(str(_crash))
    except Exception:
        pass
    sys.exit(1)

# ============================================================
# 颜色 & 字体
# ============================================================

C_BG = "#f0f0f0"
C_LOG_BG = "#1e1e1e"
C_LOG_FG = "#d4d4d4"
C_SUCCESS = "#4ec9b0"
C_WARN = "#dcdcaa"
C_ERROR = "#f44747"
C_INFO = "#569cd6"
C_NORMAL = "#d4d4d4"
C_DIM = "#888888"

F_TITLE = ("Microsoft YaHei UI", 11, "bold")
F_BODY = ("Microsoft YaHei UI", 9)
F_SMALL = ("Microsoft YaHei UI", 8)
F_LOG = ("Consolas", 9)
F_BTN = ("Microsoft YaHei UI", 10, "bold")
F_TINY = ("Microsoft YaHei UI", 7)

# 省→地级市级联数据源（用于「地域」级联筛选，按达人地域列子串匹配）
# 不带"市"字，兼容接口返回"深圳市"/"深圳"两种写法
# 省名须与 yanajiao_scraper._API_PROVINCE_CODE 的键完全一致（覆盖全部 34 省/直辖市/特别行政区）
# 城市取该省全部地级市（含自治州/地区/盟的常见简称），保证"选了省就有完整城市可选"
PROVINCE_CITIES = {
    # 直辖市
    "北京": ["北京"], "天津": ["天津"], "上海": ["上海"], "重庆": ["重庆"],
    # 河北
    "河北": ["石家庄", "唐山", "秦皇岛", "邯郸", "邢台", "保定", "张家口", "承德", "沧州", "廊坊", "衡水"],
    # 山西
    "山西": ["太原", "大同", "阳泉", "长治", "晋城", "朔州", "晋中", "运城", "忻州", "临汾", "吕梁"],
    # 内蒙古
    "内蒙古": ["呼和浩特", "包头", "乌海", "赤峰", "通辽", "鄂尔多斯", "呼伦贝尔", "巴彦淖尔", "乌兰察布", "兴安", "锡林郭勒", "阿拉善"],
    # 辽宁
    "辽宁": ["沈阳", "大连", "鞍山", "抚顺", "本溪", "丹东", "锦州", "营口", "阜新", "辽阳", "盘锦", "铁岭", "朝阳", "葫芦岛"],
    # 吉林
    "吉林": ["长春", "吉林", "四平", "辽源", "通化", "白山", "松原", "白城", "延边"],
    # 黑龙江
    "黑龙江": ["哈尔滨", "齐齐哈尔", "鸡西", "鹤岗", "双鸭山", "大庆", "伊春", "佳木斯", "七台河", "牡丹江", "黑河", "绥化", "大兴安岭"],
    # 江苏
    "江苏": ["南京", "无锡", "徐州", "常州", "苏州", "南通", "连云港", "淮安", "盐城", "扬州", "镇江", "泰州", "宿迁"],
    # 浙江
    "浙江": ["杭州", "宁波", "温州", "嘉兴", "湖州", "绍兴", "金华", "衢州", "舟山", "台州", "丽水"],
    # 安徽
    "安徽": ["合肥", "芜湖", "蚌埠", "淮南", "马鞍山", "淮北", "铜陵", "安庆", "黄山", "滁州", "阜阳", "宿州", "六安", "亳州", "池州", "宣城"],
    # 福建
    "福建": ["福州", "厦门", "莆田", "三明", "泉州", "漳州", "南平", "龙岩", "宁德"],
    # 江西
    "江西": ["南昌", "景德镇", "萍乡", "九江", "新余", "鹰潭", "赣州", "吉安", "宜春", "抚州", "上饶"],
    # 山东
    "山东": ["济南", "青岛", "淄博", "枣庄", "东营", "烟台", "潍坊", "济宁", "泰安", "威海", "日照", "临沂", "德州", "聊城", "滨州", "菏泽"],
    # 河南
    "河南": ["郑州", "开封", "洛阳", "平顶山", "安阳", "鹤壁", "新乡", "焦作", "濮阳", "许昌", "漯河", "三门峡", "南阳", "商丘", "信阳", "周口", "驻马店", "济源"],
    # 湖北
    "湖北": ["武汉", "黄石", "十堰", "宜昌", "襄阳", "鄂州", "荆门", "孝感", "荆州", "黄冈", "咸宁", "随州", "恩施"],
    # 湖南
    "湖南": ["长沙", "株洲", "湘潭", "衡阳", "邵阳", "岳阳", "常德", "张家界", "益阳", "郴州", "永州", "怀化", "娄底", "湘西"],
    # 广东
    "广东": ["广州", "韶关", "深圳", "珠海", "汕头", "佛山", "江门", "湛江", "茂名", "肇庆", "惠州", "梅州", "汕尾", "河源", "阳江", "清远", "东莞", "中山", "潮州", "揭阳", "云浮"],
    # 广西
    "广西": ["南宁", "柳州", "桂林", "梧州", "北海", "防城港", "钦州", "贵港", "玉林", "百色", "贺州", "河池", "来宾", "崇左"],
    # 海南
    "海南": ["海口", "三亚", "三沙", "儋州"],
    # 四川
    "四川": ["成都", "自贡", "攀枝花", "泸州", "德阳", "绵阳", "广元", "遂宁", "内江", "乐山", "南充", "眉山", "宜宾", "广安", "达州", "雅安", "巴中", "资阳", "阿坝", "甘孜", "凉山"],
    # 贵州
    "贵州": ["贵阳", "六盘水", "遵义", "安顺", "毕节", "铜仁", "黔西南", "黔东南", "黔南"],
    # 云南
    "云南": ["昆明", "曲靖", "玉溪", "保山", "昭通", "丽江", "普洱", "临沧", "楚雄", "红河", "文山", "西双版纳", "大理", "德宏", "怒江", "迪庆"],
    # 西藏
    "西藏": ["拉萨", "日喀则", "昌都", "林芝", "山南", "那曲", "阿里"],
    # 陕西
    "陕西": ["西安", "铜川", "宝鸡", "咸阳", "渭南", "延安", "汉中", "榆林", "安康", "商洛"],
    # 甘肃
    "甘肃": ["兰州", "嘉峪关", "金昌", "白银", "天水", "武威", "张掖", "平凉", "酒泉", "庆阳", "定西", "陇南", "临夏", "甘南"],
    # 青海
    "青海": ["西宁", "海东", "海北", "黄南", "海南", "果洛", "玉树", "海西"],
    # 宁夏
    "宁夏": ["银川", "石嘴山", "吴忠", "固原", "中卫"],
    # 新疆
    "新疆": ["乌鲁木齐", "克拉玛依", "吐鲁番", "哈密", "昌吉", "博尔塔拉", "巴音郭楞", "克孜勒苏", "伊犁", "塔城", "阿勒泰"],
    # 港澳台
    "台湾": ["台北", "高雄", "台中", "台南", "新北", "桃园", "基隆", "新竹", "嘉义"],
    "香港": ["香港"],
    "澳门": ["澳门"],
}


# ============================================================
# 多选下拉组件
# ============================================================

class MultiSelect(tk.Frame):
    """多选下拉框：点击弹出多选窗口，支持全选/清空"""

    def __init__(self, parent, options, title="", width=14, on_confirm=None, **kw):
        super().__init__(parent, **kw)
        self.options = list(options)
        self.selected = set()
        self._popup = None
        self._vars = {}
        self._on_confirm = on_confirm
        self._empty_text = "全部"
        self._btn_text = tk.StringVar(value="全部")

        if title:
            ttk.Label(self, text=title, font=F_SMALL, width=5,
                      anchor="e").pack(side="left", padx=(0, 2))

        self._btn = ttk.Button(
            self, textvariable=self._btn_text, width=width,
            command=self._toggle,
        )
        self._btn.pack(side="left")

    def _toggle(self):
        if self._popup:
            self._popup.destroy()
            self._popup = None
            return

        self._popup = tk.Toplevel(self)
        self._popup.overrideredirect(True)
        self._popup.attributes("-topmost", True)

        x = self._btn.winfo_rootx()
        y = self._btn.winfo_rooty() + self._btn.winfo_height()
        self._popup.geometry(f"+{x}+{y}")

        f = ttk.Frame(self._popup, padding=4, relief="solid", borderwidth=1)
        f.pack(fill="both", expand=True)

        # 滚动容器：选项过多时可滚动查看（修复"只能看到几个地区"）
        opt_canvas = tk.Canvas(
            f, height=min(300, 22 * len(self.options) + 10), highlightthickness=0)
        opt_sb = ttk.Scrollbar(f, orient="vertical", command=opt_canvas.yview)
        opt_inner = ttk.Frame(opt_canvas)
        opt_inner.bind(
            "<Configure>",
            lambda e: opt_canvas.configure(scrollregion=opt_canvas.bbox("all")))
        opt_canvas.create_window((0, 0), window=opt_inner, anchor="nw")
        opt_canvas.configure(yscrollcommand=opt_sb.set)
        opt_canvas.pack(side="left", fill="both", expand=True)
        opt_sb.pack(side="right", fill="y")
        opt_canvas.bind(
            "<MouseWheel>",
            lambda e: opt_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        self._vars = {}
        for opt in self.options:
            v = tk.BooleanVar(value=opt in self.selected)
            self._vars[opt] = v
            cb = ttk.Checkbutton(
                opt_inner, text=opt, variable=v,
                command=lambda o=opt: self._toggle_opt(o),
            )
            cb.pack(anchor="w", pady=1)

        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=4)
        btn_row = ttk.Frame(f)
        btn_row.pack(fill="x")
        ttk.Button(btn_row, text="全选", command=self._all).pack(side="left")
        ttk.Button(btn_row, text="清空", command=self._none).pack(side="left", padx=4)
        ttk.Button(btn_row, text="确定", command=self._confirm).pack(side="right")

    def _toggle_opt(self, opt):
        if self._vars[opt].get():
            self.selected.add(opt)
        else:
            self.selected.discard(opt)

    def _all(self):
        for o, v in self._vars.items():
            v.set(True)
            self.selected.add(o)

    def _none(self):
        for o, v in self._vars.items():
            v.set(False)
        self.selected.clear()

    def _confirm(self):
        if self._popup:
            self._popup.destroy()
            self._popup = None
        self._update_text()
        if self._on_confirm:
            self._on_confirm()

    def set_options(self, options):
        """动态重建可选项（级联筛选用）；已选但不在新范围内则清除"""
        new_opts = list(options)
        valid = set(new_opts)
        self.selected = {s for s in self.selected if s in valid}
        self.options = new_opts
        self._empty_text = "选省后可选" if not new_opts else "全部"
        self._update_text()

    def _update_text(self):
        n = len(self.selected)
        if n == 0:
            self._btn_text.set(self._empty_text)
        elif n == len(self.options):
            self._btn_text.set("全部")
        elif n <= 2:
            self._btn_text.set("、".join(sorted(self.selected)))
        else:
            self._btn_text.set(f"已选{n}项")

    def get(self):
        """返回已选列表；空或全选都返回 [] 表示不过滤"""
        n = len(self.selected)
        if n == 0 or n == len(self.options):
            return []
        return sorted(self.selected)


# ============================================================
# 级联地域选择器（省 -> 市：勾省后，右侧出现城市下拉框）
# ============================================================

class CascadeRegion(tk.Frame):
    """地域级联多选：弹窗内勾省份，该省份同一行右侧即时出现『城市 ▾』按钮，
    点击后在按钮旁边展开该省全部地级市的多选下拉（勾选即生效）。

    输出分两份：
      get_provinces() -> 选中省份列表（喂给 API 省级预筛 + 本地兜底）
      get_cities()    -> 所有已勾地级市列表（喂给本地 city 子串匹配）
    未勾任何城市时 city 为空 -> 该省整省保留；城市下拉只在该省被勾时才出现。
    """

    def __init__(self, parent, provinces, province_cities, title="", width=18, **kw):
        super().__init__(parent, **kw)
        self.provinces = list(provinces)
        self.pc = province_cities
        self._popup = None
        self._btn_text = tk.StringVar(value="全部")
        # 选择状态持久化在构造里，弹窗重建也不丢
        self._prov_vars = {p: tk.BooleanVar() for p in self.provinces}
        self._city_vars = {p: {c: tk.BooleanVar() for c in self.pc.get(p, [])}
                           for p in self.provinces}
        self._city_btns = {}       # prov -> 城市下拉按钮（仅在弹窗内存在）
        self._city_dropdowns = {}  # prov -> 城市下拉 Toplevel

        if title:
            ttk.Label(self, text=title, font=F_SMALL, width=5,
                      anchor="e").pack(side="left", padx=(0, 2))
        self._btn = ttk.Button(self, textvariable=self._btn_text, width=width,
                               command=self._toggle)
        self._btn.pack(side="left")

    def _toggle(self):
        if self._popup:
            self._close_popup()
        else:
            self._open_popup()

    def _open_popup(self):
        self._popup = tk.Toplevel(self)
        self._popup.overrideredirect(True)
        self._popup.attributes("-topmost", True)
        x = self._btn.winfo_rootx()
        y = self._btn.winfo_rooty() + self._btn.winfo_height()
        self._popup.geometry(f"+{x}+{y}")

        f = ttk.Frame(self._popup, padding=6, relief="solid", borderwidth=1)
        f.pack(fill="both", expand=True)

        cv = tk.Canvas(f, height=360, highlightthickness=0)
        sb = ttk.Scrollbar(f, orient="vertical", command=cv.yview)
        inner = ttk.Frame(cv)
        inner.bind("<Configure>",
                    lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.create_window((0, 0), window=inner, anchor="nw")
        cv.configure(yscrollcommand=sb.set)
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        cv.bind("<MouseWheel>",
                lambda e: cv.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        # 每行：省份复选框 + （勾省后）右侧『城市 ▾』下拉按钮
        self._city_btns = {}
        for p in self.provinces:
            row = ttk.Frame(inner)
            row.pack(anchor="w", fill="x", pady=1)
            cb = ttk.Checkbutton(
                row, text=p, variable=self._prov_vars[p], width=10,
                command=lambda pr=p: self._on_prov_toggle(pr))
            cb.pack(side="left", padx=(0, 8))
            if self._city_vars[p]:  # 有地级市才放下拉按钮
                btn = ttk.Button(
                    row, text=self._city_btn_text(p), width=14,
                    command=lambda pr=p: self._toggle_city_dropdown(pr))
                self._city_btns[p] = btn
                if self._prov_vars[p].get():
                    btn.pack(side="left")  # 已勾省 -> 直接显示按钮

        ttk.Separator(f, orient="horizontal").pack(fill="x", pady=4)
        btn_row = ttk.Frame(f)
        btn_row.pack(fill="x")
        ttk.Button(btn_row, text="清空", command=self._none).pack(side="left")
        ttk.Button(btn_row, text="确定", command=self._close_popup).pack(side="right")

    def _on_prov_toggle(self, prov):
        btn = self._city_btns.get(prov)
        if self._prov_vars[prov].get():
            if btn is not None:
                btn.pack(side="left")          # 勾省 -> 右侧出现城市下拉按钮
        else:
            if btn is not None:
                btn.pack_forget()
            for v in self._city_vars[prov].values():
                v.set(False)                   # 取消省 -> 清空其下城市
            self._close_city_dropdown(prov)

    def _toggle_city_dropdown(self, prov):
        if self._city_dropdowns.get(prov):
            self._close_city_dropdown(prov)
            return
        # 同一时刻只保留一个城市下拉，避免叠加
        for p in list(self._city_dropdowns.keys()):
            if p != prov:
                self._close_city_dropdown(p)
        btn = self._city_btns[prov]
        drop = tk.Toplevel(self._popup)
        drop.overrideredirect(True)
        drop.attributes("-topmost", True)
        x = btn.winfo_rootx() + btn.winfo_width()
        y = btn.winfo_rooty()
        drop.geometry(f"+{x}+{y}")

        df = ttk.Frame(drop, padding=4, relief="solid", borderwidth=1)
        df.pack(fill="both", expand=True)

        cv = tk.Canvas(df, height=min(260, 22 * len(self._city_vars[prov]) + 10),
                       highlightthickness=0)
        sb = ttk.Scrollbar(df, orient="vertical", command=cv.yview)
        inner = ttk.Frame(cv)
        inner.bind("<Configure>",
                    lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.create_window((0, 0), window=inner, anchor="nw")
        cv.configure(yscrollcommand=sb.set)
        cv.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        cv.bind("<MouseWheel>",
                lambda e: cv.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        for c, v in self._city_vars[prov].items():
            ttk.Checkbutton(
                inner, text=c, variable=v,
                command=lambda pr=prov: self._on_city_toggle(pr)).pack(anchor="w", pady=1)
        ttk.Button(df, text="确定", width=12,
                   command=lambda pr=prov: self._close_city_dropdown(pr)).pack(fill="x", pady=(2, 0))
        self._city_dropdowns[prov] = drop

    def _on_city_toggle(self, prov):
        btn = self._city_btns.get(prov)
        if btn is not None:
            btn.config(text=self._city_btn_text(prov))
        self._update_text()

    def _city_btn_text(self, prov):
        sel = [c for c, v in self._city_vars[prov].items() if v.get()]
        if not sel:
            return "全部城市 ▾"
        if len(sel) <= 2:
            return "、".join(sel) + " ▾"
        return f"已选{len(sel)}市 ▾"

    def _close_city_dropdown(self, prov):
        w = self._city_dropdowns.get(prov)
        if w:
            try:
                w.destroy()
            except Exception:
                pass
            del self._city_dropdowns[prov]

    def _none(self):
        for v in self._prov_vars.values():
            v.set(False)
        for d in self._city_vars.values():
            for v in d.values():
                v.set(False)
        for p in list(self._city_dropdowns.keys()):
            self._close_city_dropdown(p)
        for btn in self._city_btns.values():
            btn.pack_forget()
        self._update_text()

    def _close_popup(self):
        for p in list(self._city_dropdowns.keys()):
            self._close_city_dropdown(p)
        if self._popup:
            self._popup.destroy()
            self._popup = None
        self._city_btns = {}
        self._update_text()

    def _update_text(self):
        provs = [p for p in self.provinces if self._prov_vars[p].get()]
        cities = self.get_cities()
        if not provs:
            self._btn_text.set("全部")
        elif len(provs) <= 2:
            txt = "、".join(provs)
            if cities:
                txt += "·" + ("、".join(cities) if len(cities) <= 3
                              else f"{len(cities)}市")
            self._btn_text.set(txt)
        else:
            extra = f"+{len(cities)}市" if cities else ""
            self._btn_text.set(f"{len(provs)}省{extra}")

    def get_provinces(self):
        provs = [p for p in self.provinces if self._prov_vars[p].get()]
        return provs if provs else []

    def get_cities(self):
        cities = []
        for p in self.provinces:
            for c, v in self._city_vars[p].items():
                if v.get():
                    cities.append(c)
        return cities if cities else []


# ============================================================
# GUI 主类
# ============================================================

class ScraperGUI:
    """颜阿娇采集工具 GUI"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("快手达人采集工具 - 多产品版")
        self.root.geometry("1180x800")
        self.root.minsize(960, 620)
        self.root.configure(bg=C_BG)

        # --- 当前产品配置（换产品 = 换一份人群定义） ---
        self.current_product = ys.load_product()

        # --- 运行状态 ---
        self.scraper = None
        self.scraper_thread = None
        self.is_running = False
        self.start_time = None
        self.total_collected = 0
        self.files_created = []
        self.completed_tags = 0
        self.total_tags = 0
        self.current_tag = ""

        # --- 线程通信 ---
        self.msg_queue = queue.Queue()

        # --- 控件变量 ---
        self.tag_vars = {}
        self.tag_cbs = {}      # 标签 -> Checkbutton 控件（切换产品时刷新高亮）
        self.product_var = tk.StringVar(value=self.current_product.name)
        self.page_delay_var = tk.StringVar(value="1.0")
        self.tag_delay_var = tk.StringVar(value="3")
        self.page_size_var = tk.StringVar(value="20")
        self.turbo_var = tk.BooleanVar(value=False)
        self.promoter_type_var = tk.StringVar(value="全部达人")
        self.contact_filter_var = tk.StringVar(value="仅看有联系方式")
        self.status_var = tk.StringVar(value="* 就绪 - 选择标签后点击开始采集")

        # --- 联系方式提取状态 ---
        self.contact_input_var = tk.StringVar(value="")
        self.contact_delay_var = tk.StringVar(value="5")
        self.contact_max_var = tk.StringVar(value="500")
        self.contact_running = False
        self.contact_scraper_thread = None
        self.contact_start_time = None
        self.contact_extracted = 0
        self.contact_success = 0
        self.contact_skip_pause = False  # 手动跳过暂停标志

        # --- 构建 UI ---
        self._build_ui()
        self._poll_queue()

        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ========================================================
    # UI 构建
    # ========================================================

    def _build_ui(self):
        # 顶层 Notebook（标签页切换）
        nb = ttk.Notebook(self.root)
        nb.pack(fill="both", expand=True, padx=6, pady=(6, 0))

        # ======== Tab 1: 达人采集 ========
        tab1 = ttk.Frame(nb)
        nb.add(tab1, text=" 达人采集 ")

        left = ttk.LabelFrame(tab1, text="标签选择 & 设置", padding=6)
        left.pack(side="left", fill="y", padx=(0, 4))

        right = ttk.Frame(tab1)
        right.pack(side="right", fill="both", expand=True)

        self._build_left(left)
        self._build_right(right)

        # ======== Tab 2: 联系方式提取 ========
        tab2 = ttk.Frame(nb)
        nb.add(tab2, text=" 联系方式提取 ")
        self._build_contact_tab(tab2)

    def _build_left(self, parent):
        """左侧: 产品选择 + 标签选择 + 设置"""

        # --- 产品选择 ---
        pf = ttk.LabelFrame(parent, text="当前产品（人群定义）", padding=6)
        pf.pack(fill="x", pady=(0, 6))
        pr = ttk.Frame(pf)
        pr.pack(fill="x")
        ttk.Label(pr, text="产品:").pack(side="left")
        pcombo = ttk.Combobox(
            pr, textvariable=self.product_var,
            values=ys.list_products(), state="readonly", width=16,
        )
        self.product_combo = pcombo
        pcombo.pack(side="left", padx=(4, 4))
        pcombo.bind("<<ComboboxSelected>>", lambda e: self._on_product_change())
        ttk.Button(pr, text="产品管理", command=self._open_product_manager).pack(side="left")

        # --- 标签选择标题 ---
        ttk.Label(parent, text="选择采集标签", font=F_TITLE).pack(anchor="w")
        self.rec_subtitle = ttk.Label(
            parent, text="* = 推荐标签 (与产品匹配度最高)",
            font=F_SMALL, foreground=C_DIM)
        self.rec_subtitle.pack(anchor="w", pady=(0, 4))

        # --- 可滚动复选框区域 ---
        canvas_frame = ttk.Frame(parent)
        canvas_frame.pack(fill="both", expand=True, pady=(0, 4))

        canvas = tk.Canvas(canvas_frame, width=230, highlightthickness=0, bg=C_BG)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        self.scroll_frame = ttk.Frame(canvas)

        self.scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        win_id = canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        def _fit_width(e):
            canvas.itemconfig(win_id, width=e.width)
        canvas.bind("<Configure>", _fit_width)

        # 鼠标滚轮 (仅当鼠标在 canvas 区域时生效)
        def _on_wheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_wheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 创建 37 个复选框（高亮随当前产品变化）
        self._rebuild_tag_checks()

        # --- 分隔线 ---
        ttk.Separator(parent, orient="horizontal").pack(fill="x", pady=6)

        # --- 快捷按钮 ---
        btn_row = ttk.Frame(parent)
        btn_row.pack(fill="x", pady=(0, 4))
        for text, cmd in [("全选", self._sel_all), ("推荐", self._sel_rec),
                          ("清空", self._sel_none), ("反选", self._sel_inv)]:
            ttk.Button(btn_row, text=text, command=cmd).pack(side="left", padx=2)

        self.count_label = ttk.Label(parent, text="已选: 0 个", font=F_BODY)
        self.count_label.pack(anchor="w")
        self._update_count()

        # --- 分隔线 ---
        ttk.Separator(parent, orient="horizontal").pack(fill="x", pady=6)

        # --- 采集设置 ---
        sf = ttk.LabelFrame(parent, text="采集设置", padding=6)
        sf.pack(fill="x")

        # 达人类型
        r0 = ttk.Frame(sf)
        r0.pack(fill="x", pady=2)
        ttk.Label(r0, text="达人类型:", font=F_BODY).pack(side="left")
        ttk.Combobox(r0, textvariable=self.promoter_type_var, width=10,
                      values=["全部达人", "直播达人", "视频达人"],
                      state="readonly", font=F_BODY).pack(side="left", padx=4)

        # 联系方式
        r0b = ttk.Frame(sf)
        r0b.pack(fill="x", pady=2)
        ttk.Label(r0b, text="联系方式:", font=F_BODY).pack(side="left")
        ttk.Combobox(r0b, textvariable=self.contact_filter_var, width=14,
                     values=["仅看有联系方式", "仅看无联系方式", "全部达人"],
                     state="readonly", font=F_BODY).pack(side="left", padx=4)

        # --- 筛选条件（v11.5: 平台先筛后拉；不限=全部） ---
        ttk.Separator(sf, orient="horizontal").pack(fill="x", pady=4)
        ttk.Label(sf, text="筛选条件（平台先筛后拉）", font=F_SMALL,
                  foreground=C_DIM).pack(anchor="w", pady=(0, 2))

        self.filter_gender = MultiSelect(sf, ["男", "女"], "性别:", width=13)
        self.filter_gender.pack(fill="x", pady=1)

        # 地域（省→市 级联）：省份取自 API 省份码映射（34省全量），城市按其下地级市
        # 勾省后，该行右侧出现『城市 ▾』下拉按钮，点击在旁边展开该省全部城市多选
        # 平台 API 仅支持省级预筛；地级市需在『采集粉丝画像=开』时按达人地域列本地匹配
        regions = list(ys.YanajiaoScraper._API_PROVINCE_CODE.keys())
        self.filter_region = CascadeRegion(sf, regions, PROVINCE_CITIES,
                                           "地域:", width=18)
        self.filter_region.pack(fill="x", pady=1)
        ttk.Label(
            sf, text="勾省份→右侧出现『城市▾』下拉选具体地级市；地级市需开启『采集粉丝画像』才生效",
            font=F_TINY, foreground=C_DIM).pack(anchor="w", pady=(0, 1))

        ages = ["18~23岁", "24~30岁", "31~40岁", "41~50岁", "50岁以上"]
        self.filter_fan_age = MultiSelect(sf, ages, "年龄:", width=13)
        self.filter_fan_age.pack(fill="x", pady=1)

        self.filter_fan_gender = MultiSelect(
            sf, ["女性为主", "男性为主"], "粉性:", width=13)
        self.filter_fan_gender.pack(fill="x", pady=1)

        cities = ["一线城市", "新一线城市", "二线城市",
                  "三线城市", "四线城市", "五线城市"]
        self.filter_fan_city = MultiSelect(sf, cities, "城级:", width=13)
        self.filter_fan_city.pack(fill="x", pady=1)

        # 分隔
        ttk.Separator(sf, orient="horizontal").pack(fill="x", pady=4)

        # 页间延迟
        r1 = ttk.Frame(sf)
        r1.pack(fill="x", pady=2)
        ttk.Label(r1, text="页间延迟:", font=F_BODY).pack(side="left")
        ttk.Spinbox(r1, from_=0.5, to=5.0, increment=0.5,
                     textvariable=self.page_delay_var, width=5,
                     font=F_BODY).pack(side="left", padx=4)
        ttk.Label(r1, text="秒", font=F_SMALL).pack(side="left")

        # 标签间延迟
        r2 = ttk.Frame(sf)
        r2.pack(fill="x", pady=2)
        ttk.Label(r2, text="标签间延迟:", font=F_BODY).pack(side="left")
        ttk.Spinbox(r2, from_=1, to=30, increment=1,
                     textvariable=self.tag_delay_var, width=5,
                     font=F_BODY).pack(side="left", padx=4)
        ttk.Label(r2, text="秒", font=F_SMALL).pack(side="left")

        # 每页数量
        r3 = ttk.Frame(sf)
        r3.pack(fill="x", pady=2)
        ttk.Label(r3, text="每页拉取:", font=F_BODY).pack(side="left")
        ttk.Combobox(r3, textvariable=self.page_size_var, width=5,
                      values=["20", "50"], state="readonly",
                      font=F_BODY).pack(side="left", padx=4)
        ttk.Label(r3, text="条/页", font=F_SMALL).pack(side="left")

        # 极速模式
        ttk.Checkbutton(sf, text="极速模式 (0.5秒/页 + 50条/页)",
                        variable=self.turbo_var,
                        command=self._on_turbo).pack(anchor="w", pady=(6, 0))

        ttk.Label(sf, text="极速模式可能触发限流\n但限流后自动退避重试, 不影响结果",
                  font=F_SMALL, foreground=C_DIM, justify="left").pack(anchor="w", pady=(4, 0))

    def _build_right(self, parent):
        """右侧: 控制按钮 + 进度 + 日志 + 状态栏"""        # --- 控制按钮 ---
        ctrl = ttk.Frame(parent)
        ctrl.pack(fill="x", pady=(0, 4))

        self.btn_start = ttk.Button(ctrl, text=">>  开始采集", command=self._on_start)
        self.btn_start.pack(side="left", padx=(0, 4))

        self.btn_stop = ttk.Button(ctrl, text="[]  停止", command=self._on_stop, state="disabled")
        self.btn_stop.pack(side="left", padx=4)

        ttk.Button(ctrl, text="打开文件夹", command=self._on_open).pack(side="left", padx=4)
        ttk.Button(ctrl, text="清空日志", command=self._clear_log).pack(side="left", padx=4)

        # --- 进度面板 ---
        pf = ttk.LabelFrame(parent, text="采集进度", padding=6)
        pf.pack(fill="x", pady=(0, 4))

        # 当前标签
        r1 = ttk.Frame(pf)
        r1.pack(fill="x", pady=(0, 2))
        self.lbl_cur = ttk.Label(r1, text="当前: -", font=F_BODY)
        self.lbl_cur.pack(side="left")
        self.lbl_cur2 = ttk.Label(r1, text="", font=F_SMALL, foreground=C_DIM)
        self.lbl_cur2.pack(side="right")

        self.pb_tag = ttk.Progressbar(pf, mode="determinate")
        self.pb_tag.pack(fill="x", pady=(0, 6))

        # 总进度
        r2 = ttk.Frame(pf)
        r2.pack(fill="x", pady=(0, 2))
        self.lbl_all = ttk.Label(r2, text="总进度: 0/0", font=F_BODY)
        self.lbl_all.pack(side="left")
        self.lbl_all2 = ttk.Label(r2, text="", font=F_SMALL, foreground=C_DIM)
        self.lbl_all2.pack(side="right")

        self.pb_all = ttk.Progressbar(pf, mode="determinate")
        self.pb_all.pack(fill="x")

        # 统计
        self.lbl_stats = ttk.Label(pf, text="", font=F_SMALL, foreground=C_DIM)
        self.lbl_stats.pack(anchor="w", pady=(4, 0))

        # --- 日志面板 ---
        lf = ttk.LabelFrame(parent, text="运行日志", padding=2)
        lf.pack(fill="both", expand=True, pady=(0, 4))

        self.log_text = tk.Text(
            lf, wrap="word", font=F_LOG,
            bg=C_LOG_BG, fg=C_LOG_FG,
            insertbackground=C_LOG_FG,
            selectbackground="#264f78",
            relief="flat", padx=8, pady=4,
        )
        log_sb = ttk.Scrollbar(lf, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_sb.set)

        # 日志颜色
        self.log_text.tag_configure("success", foreground=C_SUCCESS)
        self.log_text.tag_configure("warning", foreground=C_WARN)
        self.log_text.tag_configure("error", foreground=C_ERROR)
        self.log_text.tag_configure("info", foreground=C_INFO)
        self.log_text.tag_configure("normal", foreground=C_NORMAL)

        self.log_text.pack(side="left", fill="both", expand=True)
        log_sb.pack(side="right", fill="y")

        # --- 状态栏 ---
        self.lbl_status = ttk.Label(
            parent, textvariable=self.status_var,
            font=F_BODY, relief="sunken", anchor="w", padding=(8, 3),
        )
        self.lbl_status.pack(fill="x")

    def _build_contact_tab(self, parent):
        """联系方式提取标签页"""
        # 左侧：文件选择 + 设置
        left = ttk.LabelFrame(parent, text="输入 & 设置", padding=6)
        left.pack(side="left", fill="y", padx=(0, 4))

        # 文件选择
        ttk.Label(left, text="选择采集结果Excel", font=F_TITLE).pack(anchor="w")
        ttk.Label(left, text="会提取文件中每个达人的联系方式",
                  font=F_SMALL, foreground=C_DIM).pack(anchor="w", pady=(0, 4))

        fr = ttk.Frame(left)
        fr.pack(fill="x", pady=(0, 4))
        ttk.Entry(fr, textvariable=self.contact_input_var, font=F_BODY,
                  state="readonly").pack(side="left", fill="x", expand=True)
        ttk.Button(fr, text="浏览", command=self._on_contact_browse
                   ).pack(side="right", padx=(4, 0))

        # 快捷选择
        ttk.Button(left, text="自动选最新Excel", command=self._on_contact_latest
                   ).pack(anchor="w", pady=(0, 6))

        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=4)

        # 设置
        ttk.Label(left, text="提取设置", font=F_TITLE).pack(anchor="w", pady=(4, 0))

        r1 = ttk.Frame(left)
        r1.pack(fill="x", pady=4)
        ttk.Label(r1, text="每人间隔:", font=F_BODY).pack(side="left")
        ttk.Spinbox(r1, from_=3, to=30, increment=1,
                     textvariable=self.contact_delay_var, width=5,
                     font=F_BODY).pack(side="left", padx=4)
        ttk.Label(r1, text="秒", font=F_SMALL).pack(side="left")

        r2 = ttk.Frame(left)
        r2.pack(fill="x", pady=4)
        ttk.Label(r2, text="每次最多提取:", font=F_BODY).pack(side="left")
        ttk.Spinbox(r2, from_=5, to=200, increment=5,
                     textvariable=self.contact_max_var, width=5,
                     font=F_BODY).pack(side="left", padx=4)
        ttk.Label(r2, text="人", font=F_SMALL).pack(side="left")

        ttk.Label(left, text="建议每次不超过50人，避免触发风控",
                  font=F_SMALL, foreground=C_DIM).pack(anchor="w", pady=(0, 4))

        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=8)

        # 注意事项
        ttk.Label(left, text="⚠️ 使用说明", font=F_TITLE).pack(anchor="w")
        notes = [
            "1. 确保已登录快手（与达人采集共用登录态）",
            "2. 浏览器窗口会打开，请勿手动操作",
            "3. 每20人会自动暂停15秒防限流",
            "4. 进度自动保存，中途退出可恢复",
        ]
        for n in notes:
            ttk.Label(left, text=n, font=F_SMALL,
                      foreground=C_DIM, wraplength=220).pack(anchor="w", pady=1)

        # 右侧：控制 + 进度 + 日志
        right = ttk.Frame(parent)
        right.pack(side="right", fill="both", expand=True)

        # 控制按钮
        ctrl = ttk.Frame(right)
        ctrl.pack(fill="x", pady=(0, 4))

        self.btn_contact_start = ttk.Button(ctrl, text=">>  开始提取联系方式",
                                            command=self._on_contact_start)
        self.btn_contact_start.pack(side="left", padx=(0, 4))

        self.btn_contact_stop = ttk.Button(ctrl, text="[]  停止",
                                           command=self._on_contact_stop,
                                           state="disabled")
        self.btn_contact_stop.pack(side="left", padx=4)

        self.btn_contact_skip = ttk.Button(ctrl, text="⏹  停止暂停",
                                           command=self._on_contact_skip_pause,
                                           state="disabled")
        self.btn_contact_skip.pack(side="left", padx=4)

        ttk.Button(ctrl, text="打开文件夹",
                   command=lambda: os.startfile(str(ys.Config.OUTPUT_DIR))
                   ).pack(side="left", padx=4)

        ttk.Separator(ctrl, orient="vertical").pack(side="left", fill="y", padx=6)
        self.btn_contact_copy = ttk.Button(ctrl, text="📋  复制到表格",
                                            command=self._on_contact_copy)
        self.btn_contact_copy.pack(side="left", padx=4)

        # 进度
        pf = ttk.LabelFrame(right, text="提取进度", padding=6)
        pf.pack(fill="x", pady=(0, 4))

        self.lbl_contact_progress = ttk.Label(pf, text="等待开始...", font=F_BODY)
        self.lbl_contact_progress.pack(anchor="w")

        self.pb_contact = ttk.Progressbar(pf, mode="determinate")
        self.pb_contact.pack(fill="x", pady=(4, 0))

    def _on_contact_browse(self):
        """浏览选择Excel文件"""
        fp = filedialog.askopenfilename(
            title="选择采集结果Excel",
            initialdir=str(ys.Config.OUTPUT_DIR),
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if fp:
            self.contact_input_var.set(fp)
            self._log(f"已选择: {Path(fp).name}", "info")

    def _on_contact_latest(self):
        """自动选择最新Excel"""
        excels = list(ys.Config.OUTPUT_DIR.glob("*_达人_*_*.xlsx"))
        excels = [f for f in excels if "联系方式" not in f.stem]
        if not excels:
            messagebox.showwarning("提示", "采集结果中暂无Excel文件")
            return
        excels.sort(key=lambda f: f.stat().st_mtime, reverse=True)
        self.contact_input_var.set(str(excels[0]))
        self._log(f"自动选择最新: {excels[0].name}", "info")

    def _on_contact_start(self):
        """开始提取联系方式"""
        input_file = self.contact_input_var.get()
        if not input_file:
            messagebox.showwarning("提示", "请先选择采集结果Excel!")
            return
        if not Path(input_file).exists():
            messagebox.showerror("错误", f"文件不存在:\n{input_file}")
            return
        if self.contact_running:
            return

        try:
            delay = int(self.contact_delay_var.get())
            max_count = int(self.contact_max_var.get())
        except ValueError:
            messagebox.showerror("错误", "请输入有效的数字!")
            return

        self.contact_running = True
        self.contact_extracted = 0
        self.contact_success = 0
        self.contact_skip_pause = False
        self.contact_start_time = time.time()

        self.btn_contact_start.config(state="disabled")
        self.btn_contact_stop.config(state="normal")
        self.btn_contact_copy.config(state="disabled")
        self.pb_contact["value"] = 0
        self.lbl_contact_progress.config(text="启动浏览器中...")

        self._log("=" * 50, "normal")
        self._log(f">> 开始提取联系方式: {Path(input_file).name}", "info")
        self._log(f"   间隔: {delay}秒 | 上限: {max_count}人", "normal")
        self._log("=" * 50, "normal")

        # 设置参数
        ec.Config.PAGE_DELAY = delay
        ec.Config.MAX_PER_SESSION = max_count

        # 启动后台线程
        self.contact_scraper_thread = threading.Thread(
            target=self._run_contact_extract, args=(input_file, max_count),
            daemon=True
        )
        self.contact_scraper_thread.start()
        self._tick_timer_contact()

    def _on_contact_stop(self):
        """停止提取"""
        self.contact_running = False
        self._log("🛑 收到停止信号...", "warning")
        self.btn_contact_stop.config(state="disabled")
        self.btn_contact_skip.config(state="disabled")
        self.btn_contact_copy.config(state="disabled")

    def _on_contact_skip_pause(self):
        """手动停止暂停，立即恢复提取"""
        self.contact_skip_pause = True
        self._log("⏹  收到停止暂停信号，立即恢复...", "info")
        self.btn_contact_skip.config(state="disabled")

    def _on_contact_copy(self):
        """把已提取到的联系方式原子合并回原表格。

        这样下次打开同一表格提取时，已有联系方式的达人会被自动跳过，
        不浪费每天的查看名额。可随时点击（提取进行中按钮禁用）。
        """
        input_file = self.contact_input_var.get()
        if not input_file or not Path(input_file).exists():
            self._log("请先选择采集结果Excel", "warning")
            return
        from extract_contacts import _resolve_output_file, finalize_to_original
        out = _resolve_output_file(input_file)
        if not Path(out).exists():
            self._log("还没有可复制的联系方式，请先提取", "warning")
            return
        try:
            finalize_to_original(input_file, out)
            self._log(f"✅ 已把已提取的联系方式复制到原表格: {Path(input_file).name}", "success")
            self._log("   下次打开此表格提取时，已有联系方式会直接跳过，不浪费名额", "normal")
        except PermissionError:
            self._log("❌ 复制失败：原表格正被 Excel 打开，请先关闭后再点“复制到表格”", "error")
        except Exception as e:
            self._log(f"❌ 复制到表格失败: {e}", "error")

    def _run_contact_extract(self, input_file, max_count):
        """后台线程：批量提取联系方式"""
        from extract_contacts import batch_extract, extract_one, Config as EcConfig
        from playwright.sync_api import sync_playwright

        EcConfig.PAGE_DELAY = int(self.contact_delay_var.get())
        EcConfig.MAX_PER_SESSION = max_count

        try:
            results = []

            # 读取Excel
            from openpyxl import load_workbook
            wb = load_workbook(input_file, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            wb.close()

            if not rows:
                self.msg_queue.put(("contact_error", "Excel为空"))
                return

            header = [str(h).strip() if h else "" for h in rows[0]]
            data_rows = rows[1:]

            # 找列
            pid_col = None
            name_col = None
            phone_col = None
            wechat_col = None
            for i, h in enumerate(header):
                hl = h.lower()
                if "快手id" in hl or "promoter" in hl or h == "快手ID":
                    pid_col = i
                if "达人名称" in hl or "昵称" in hl or h == "达人名称":
                    name_col = i
                if h == "手机号":
                    phone_col = i
                if h == "微信号":
                    wechat_col = i

            if pid_col is None:
                self.msg_queue.put(("contact_error", "未找到快手ID列"))
                return

            # 原文件保持只读；结果写入独立的“_联系方式”文件，任何异常都不损坏原始采集数据
            from extract_contacts import _init_output_file, load_work_extracted_pids
            output_file = _init_output_file(input_file)
            # 加载工作文件里【已提取】的达人ID集合：即便还没点“复制到表格”合并回原文件，
            # 这些达人也应直接跳过，不重复消耗每日名额（断点续提的持久依据）。
            work_extracted_pids = load_work_extracted_pids(output_file)
            if work_extracted_pids:
                self.msg_queue.put(("contact_log", f"   📌 工作文件中已提取: {len(work_extracted_pids)} 人 (未合并回原文件也会跳过)"))

            # 过滤：跳过已有联系方式的达人（原文件列 或 工作文件已提取记录 任一命中即跳过）
            rows_to_extract = []  # (原始索引, 行数据)
            already_has = 0
            for idx, row in enumerate(data_rows):
                has_contact = False
                if phone_col is not None and row[phone_col]:
                    val = str(row[phone_col]).strip()
                    if val and val not in ("None", ""):
                        has_contact = True
                if wechat_col is not None and row[wechat_col]:
                    val = str(row[wechat_col]).strip()
                    if val and val not in ("None", ""):
                        has_contact = True
                pid_val = str(row[pid_col]).strip() if pid_col is not None and row[pid_col] else ""
                if pid_val and pid_val in work_extracted_pids:
                    has_contact = True
                if has_contact:
                    already_has += 1
                else:
                    rows_to_extract.append((idx, row))

            total = len(data_rows)
            need_extract = len(rows_to_extract)
            effective = min(max_count, need_extract)

            if already_has > 0:
                self.msg_queue.put(("contact_log", f"   ⏭ 已有联系方式: {already_has} 人 (自动跳过)"))
            self.msg_queue.put(("contact_log", f"   🎯 待提取: {need_extract} 人 | 本次: {effective} 人"))

            if need_extract == 0:
                self.msg_queue.put(("contact_log", "✅ 所有达人联系方式已齐全！"))
                self.msg_queue.put(("contact_done", [], output_file))
                return

            self.msg_queue.put(("contact_total", effective))
            self.pb_contact["maximum"] = effective

            # 启动浏览器（辅助函数，GUI内联以共享消息队列）
            def _open_contact_browser():
                self.msg_queue.put(("contact_log", "🚀 启动浏览器..."))
                pp = sync_playwright().start()
                cctx = pp.chromium.launch_persistent_context(
                    user_data_dir=str(EcConfig.USER_DATA),
                    headless=False,
                    args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
                    viewport={"width": 1440, "height": 900},
                    locale="zh-CN",
                )
                ppage = cctx.new_page()
                ppage.goto(EcConfig.DAREN_SQUARE_URL, timeout=60000, wait_until="domcontentloaded")
                time.sleep(5)
                if "login" in ppage.url.lower():
                    self.msg_queue.put(("contact_log", "⚠️  需要扫码登录快手!"))
                return pp, cctx, ppage

            def _close_contact_browser(pp, cctx):
                try:
                    cctx.close()
                except Exception:
                    pass
                try:
                    pp.stop()
                except Exception:
                    pass

            p, ctx, page = _open_contact_browser()

            # 逐个提取（while循环，支持自动暂停恢复）
            consecutive_empty = 0
            pause_count = 0
            stopped_by_user = False
            i = 0
            while i < effective:
                if not self.contact_running:
                    stopped_by_user = True
                    break

                orig_idx, row = rows_to_extract[i]
                pid = str(row[pid_col]) if row[pid_col] else ""
                name = str(row[name_col]) if name_col is not None and len(row) > name_col and row[name_col] else ""

                if not pid or pid.upper() == "NONE":
                    i += 1
                    continue

                self.msg_queue.put(("contact_progress", i, effective, name, pid))

                result = extract_one(page, pid, name)

                if result["phone"] or result["wechat"]:
                    self.contact_success += 1
                    consecutive_empty = 0
                else:
                    consecutive_empty += 1
                self.contact_extracted += 1
                results.append(result)

                # 保存进度：JSON 每步存（快速可靠），Excel 每5步批量写（防卡锁）
                from extract_contacts import _save_progress, _save_final_excel, finalize_to_original
                _save_progress(results, input_file)
                if i % 5 == 0 or i == effective - 1 or result["phone"] or result["wechat"]:
                    try:
                        _save_final_excel(results, output_file)
                    except Exception as e:
                        self.msg_queue.put(("contact_log", f"⚠️ Excel保存失败（JSON已存）: {e}"))

                i += 1

                # 检测：连续N个无联系方式 → 自动暂停恢复
                if consecutive_empty >= EcConfig.AUTO_PAUSE_CONSECUTIVE_EMPTY:
                    pause_count += 1
                    self.msg_queue.put(("contact_log", ""))
                    self.msg_queue.put(("contact_log", "⚠️" * 20))
                    self.msg_queue.put(("contact_log", f"⚠️  连续 {consecutive_empty} 个达人无联系方式！"))
                    self.msg_queue.put(("contact_log", "⚠️  疑似触达快手每日查询上限，触发自动暂停..."))
                    self.msg_queue.put(("contact_log", "⚠️" * 20))

                    # 先保存已有结果
                    _save_final_excel(results, output_file)

                    # 不关浏览器！保持登录会话存活，避免恢复时需要人工扫码
                    page.goto(EcConfig.DAREN_SQUARE_URL, timeout=30000, wait_until="domcontentloaded")
                    time.sleep(2)

                    wait_minutes = EcConfig.AUTO_PAUSE_MINUTES
                    self.msg_queue.put(("contact_log", f"⏰ 暂停 {wait_minutes} 分钟（浏览器保持打开，无需重新登录）"))
                    self.msg_queue.put(("contact_log", f"   (已提取 {len(results)} 人，成功 {self.contact_success} 人，第 {pause_count} 次暂停)"))
                    self.msg_queue.put(("contact_log", f"   💡 可点击「停止暂停」按钮立即恢复"))
                    self.lbl_contact_progress.config(text=f"⏰ 暂停中... 剩余约 {wait_minutes} 分钟（可点停止暂停）")

                    # 启用跳过暂停按钮
                    self.contact_skip_pause = False
                    self.msg_queue.put(("contact_pause_started",))

                    # 分段等待，每5秒检查一次跳过标志（更快响应）
                    wait_seconds = wait_minutes * 60
                    skipped = False
                    while wait_seconds > 0 and self.contact_running:
                        if self.contact_skip_pause:
                            skipped = True
                            break
                        chunk = min(5, wait_seconds)
                        time.sleep(chunk)
                        wait_seconds -= chunk
                        if wait_seconds > 0:
                            rem_m = wait_seconds // 60
                            rem_s = wait_seconds % 60
                            self.lbl_contact_progress.config(text=f"⏰ 暂停中... 剩余约 {rem_m}分{rem_s}秒（可点停止）")
                            self.msg_queue.put(("contact_log", f"   ⏳ 剩余约 {rem_m} 分 {rem_s} 秒..."))

                    # 禁用跳过暂停按钮
                    self.msg_queue.put(("contact_pause_ended",))

                    if not self.contact_running:
                        stopped_by_user = True
                        break

                    if skipped:
                        self.msg_queue.put(("contact_log", "⏹  用户手动停止暂停，立即恢复提取"))
                    else:
                        self.msg_queue.put(("contact_log", f"⏰ 等待完成，自动恢复"))

                    # 快速验证登录态（导航到广场确认，无需重开浏览器）
                    self.msg_queue.put(("contact_log", ""))
                    self.msg_queue.put(("contact_log", "=" * 40))
                    self.msg_queue.put(("contact_log", f"🔄 第 {pause_count} 次恢复：验证登录态..."))
                    self.lbl_contact_progress.config(text=f"验证登录态...")

                    page.goto(EcConfig.DAREN_SQUARE_URL, timeout=30000, wait_until="domcontentloaded")
                    time.sleep(3)

                    if "login" in page.url.lower():
                        self.msg_queue.put(("contact_log", "⚠️  登录态已过期，等待恢复（30秒自动检测）..."))
                        self.lbl_contact_progress.config(text=f"登录态过期，等待恢复...")
                        logged_in = False
                        for _ in range(15):
                            time.sleep(2)
                            if "login" not in page.url.lower():
                                self.msg_queue.put(("contact_log", "🔓 登录已恢复！"))
                                logged_in = True
                                break
                        if not logged_in:
                            self.msg_queue.put(("contact_log", "⚠️  需要手动扫码，请查看浏览器窗口..."))
                            self.lbl_contact_progress.config(text=f"需要扫码登录（60秒等待）")
                            for _ in range(60):
                                if not self.contact_running:
                                    stopped_by_user = True
                                    break
                                time.sleep(2)
                                if "login" not in page.url.lower():
                                    self.msg_queue.put(("contact_log", "🔓 登录成功！"))
                                    logged_in = True
                                    break
                        if not logged_in:
                            self.msg_queue.put(("contact_log", "❌ 登录超时，停止提取"))
                            stopped_by_user = True
                            break

                    consecutive_empty = 0
                    self.msg_queue.put(("contact_log", f"✅ 已恢复，从第 {i+1}/{effective} 个达人继续"))
                    self.lbl_contact_progress.config(text=f"已恢复，继续提取...")
                    continue

                # 频率控制（正常模式）
                if i < effective and self.contact_running:
                    time.sleep(EcConfig.PAGE_DELAY)
                    if i % EcConfig.LONG_PAUSE_EVERY == 0 and i > 0:
                        self.msg_queue.put(("contact_log", f"  🫁 暂停 {EcConfig.LONG_PAUSE_DURATION} 秒..."))
                        time.sleep(EcConfig.LONG_PAUSE_DURATION)

            # 输出
            if stopped_by_user:
                self.msg_queue.put(("contact_log", ""))
                self.msg_queue.put(("contact_log", "🛑 用户手动停止"))
                self.msg_queue.put(("contact_log", f"   已保存 {self.contact_extracted} 人（成功 {self.contact_success} 人）到结果文件"))
                _save_final_excel(results, output_file)
                # 最终：原子合并联系方式回原文件（原文件即最终成品）
                finalize_to_original(input_file, output_file)
                output_file = input_file
                self.msg_queue.put(("contact_done", results, output_file))
            else:
                _save_final_excel(results, output_file)
                # 最终：原子合并联系方式回原文件（原文件即最终成品）
                finalize_to_original(input_file, output_file)
                output_file = input_file
                self.msg_queue.put(("contact_done", results, output_file))

            _close_contact_browser(p, ctx)

        except Exception as e:
            import traceback
            traceback.print_exc()
            self.msg_queue.put(("contact_error", str(e)))

    def _tick_timer_contact(self):
        """联系方式提取的计时器"""
        if self.contact_running:
            self.root.after(1000, self._tick_timer_contact)

    # ========================================================
    # 标签选择操作
    # ========================================================

    def _get_tags(self) -> list:
        return [t for t in ys.Config.ALL_TAGS if self.tag_vars[t].get()]

    def _rebuild_tag_checks(self):
        """按当前产品重建标签复选框（推荐标签高亮 + 默认勾选）"""
        # 清空旧控件
        for cb in self.tag_cbs.values():
            cb.destroy()
        self.tag_cbs.clear()
        self.tag_vars.clear()

        rec_set = set(self.current_product.recommended_tags or [])
        for tag in ys.Config.ALL_TAGS:
            is_rec = tag in rec_set
            prefix = "* " if is_rec else "   "
            var = tk.BooleanVar(value=is_rec)
            self.tag_vars[tag] = var
            cb = ttk.Checkbutton(
                self.scroll_frame,
                text=f"{prefix}{tag}",
                variable=var,
                command=self._update_count,
            )
            cb.pack(anchor="w", padx=4, pady=1)
            self.tag_cbs[tag] = cb
        self._update_count()

    def _refresh_tag_highlights(self):
        """切换产品后，只刷新高亮文案（不重建控件）"""
        rec_set = set(self.current_product.recommended_tags or [])
        for tag, cb in self.tag_cbs.items():
            prefix = "* " if tag in rec_set else "   "
            cb.config(text=f"{prefix}{tag}")

    def _on_product_change(self):
        """下拉切换产品：重新载入人群定义并刷新高亮 + 默认勾选推荐"""
        name = self.product_var.get()
        if not name:
            return
        self.current_product = ys.load_product(name)
        self.root.title(f"快手达人采集工具 - {self.current_product.name}")
        self.rec_subtitle.config(
            text=f"* = 推荐标签 (与「{self.current_product.name}」匹配度最高)")
        self._refresh_tag_highlights()
        # 切换产品后，自动勾选该产品的推荐标签，方便直接开采
        self._sel_rec()
        self._log(f">> 已切换产品: {self.current_product.name}（关联列: "
                  f"{self.current_product.assoc_field}）", "info")

    def _open_product_manager(self):
        """产品管理窗口：可视化编辑/新建/删除产品的人群定义"""
        ProductManager(self.root, self.current_product.name, self._on_product_saved)

    def _on_product_saved(self, name: str):
        """产品保存/删除后回调：刷新下拉框并切到该产品"""
        self.product_var.set(name)
        self.current_product = ys.load_product(name)
        self._refresh_product_combos()
        self._refresh_tag_highlights()
        self._sel_rec()

    def _refresh_product_combos(self):
        """刷新产品下拉框的候选列表（可能新增/删除了产品）"""
        try:
            self.product_combo["values"] = ys.list_products()
        except Exception:
            pass

    def _update_count(self):
        if not hasattr(self, "count_label"):
            return
        n = len(self._get_tags())
        self.count_label.config(text=f"已选: {n} 个")

    def _sel_all(self):
        for v in self.tag_vars.values():
            v.set(True)
        self._update_count()

    def _sel_rec(self):
        rec = set(self.current_product.recommended_tags or [])
        for t, v in self.tag_vars.items():
            v.set(t in rec)
        self._update_count()

    def _sel_none(self):
        for v in self.tag_vars.values():
            v.set(False)
        self._update_count()

    def _sel_inv(self):
        for v in self.tag_vars.values():
            v.set(not v.get())
        self._update_count()

    def _on_turbo(self):
        if self.turbo_var.get():
            self.page_delay_var.set("0.5")
            self.tag_delay_var.set("1")
            self.page_size_var.set("50")
            self._log("极速模式已开启: 页间0.5秒, 标签间1秒, 每页50条", "info")
        else:
            self.page_delay_var.set("1.0")
            self.tag_delay_var.set("3")
            self.page_size_var.set("20")
            self._log("极速模式已关闭", "normal")

    # ========================================================
    # 采集控制
    # ========================================================

    def _on_start(self):
        """开始采集"""
        tags = self._get_tags()
        if not tags:
            messagebox.showwarning("提示", "请至少选择一个标签!")
            return
        if self.is_running:
            return

        if len(tags) > 15:
            if not messagebox.askyesno("确认",
                f"将采集 {len(tags)} 个标签, 可能需要较长时间. 继续?"):
                return

        # 重置状态
        self.is_running = True
        self.total_collected = 0
        self.files_created = []
        self.completed_tags = 0
        self.total_tags = len(tags)
        self.start_time = time.time()

        # 更新 UI
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self._set_status("running", f">> 采集中... 0/{self.total_tags} 标签")

        self.pb_all["maximum"] = self.total_tags
        self.pb_all["value"] = 0
        self.pb_tag["value"] = 0
        self.lbl_all.config(text=f"总进度: 0/{self.total_tags}")
        self.lbl_cur.config(text="当前: -")

        # 收集筛选条件
        self.filter_config = {
            "gender": self.filter_gender.get(),
            "region": self.filter_region.get_provinces(),
            "city": self.filter_region.get_cities(),
            "fan_age": self.filter_fan_age.get(),
            "fan_gender": self.filter_fan_gender.get(),
            "fan_city": self.filter_fan_city.get(),
        }
        active = {k: v for k, v in self.filter_config.items() if v}
        filter_desc = " | ".join(f"{k}={','.join(v)}" for k, v in active.items()) if active else "无"

        self._log("=" * 50, "normal")
        self._log(f">> 开始采集 {len(tags)} 个标签", "info")
        self._log(f"   达人类型: {self.promoter_type_var.get()}", "normal")
        self._log(f"   联系方式: {self.contact_filter_var.get()}", "normal")
        self._log(f"   筛选条件: {filter_desc}", "normal")
        self._log(f"   页间延迟: {self.page_delay_var.get()}秒 | "
                  f"标签间: {self.tag_delay_var.get()}秒 | "
                  f"每页: {self.page_size_var.get()}条", "normal")
        self._log("=" * 50, "normal")

        # 启动采集线程
        self.scraper_thread = threading.Thread(
            target=self._run_scraper, args=(tags,), daemon=True
        )
        self.scraper_thread.start()
        self._tick_timer()

    def _on_stop(self):
        """停止采集"""
        if self.scraper and self.is_running:
            self.scraper.stop()
            self._set_status("stopped", "[] 正在停止...")
            self.btn_stop.config(state="disabled")

    def _run_scraper(self, tags: list):
        """采集线程: 在后台运行, 通过队列与 GUI 通信"""
        # Monkey-patch log 函数
        original_log = ys.log

        def gui_log(msg, end="\n"):
            self.msg_queue.put(("log", msg))

        ys.log = gui_log

        try:
            self.scraper = ys.YanajiaoScraper(self.current_product.name)
            # 设置回调 -> 通过队列传给 GUI
            self.scraper._on_progress = lambda tag, collected, total, page: \
                self.msg_queue.put(("progress", tag, collected, total, page))
            self.scraper._on_tag_done = lambda tag, count, filepath: \
                self.msg_queue.put(("tag_done", tag, count, filepath))
            self.scraper._on_all_done = lambda results, files, elapsed: \
                self.msg_queue.put(("all_done", results, files, elapsed))

            # 应用设置
            self.scraper.page_delay = float(self.page_delay_var.get())
            self.scraper.tag_delay = int(self.tag_delay_var.get())
            self.scraper.page_size = int(self.page_size_var.get())
            # 达人类型
            type_map = {"全部达人": 0, "直播达人": 1, "视频达人": 2}
            self.scraper.promoter_type = type_map.get(
                self.promoter_type_var.get(), 0)
            # 联系方式
            contact_map = {"仅看有联系方式": "with",
                           "仅看无联系方式": "without",
                           "全部达人": "all"}
            self.scraper.contact_filter = contact_map.get(
                self.contact_filter_var.get(), "with")
            # 筛选条件
            self.scraper.filter_config = getattr(self, "filter_config", {})

            # 执行
            self.scraper.run(tags)

        except Exception as e:
            self.msg_queue.put(("error", str(e)))
            import traceback
            traceback.print_exc()
        finally:
            ys.log = original_log
            self.msg_queue.put(("thread_end",))

    def _on_open(self):
        """打开输出文件夹"""
        os.startfile(str(ys.Config.OUTPUT_DIR))

    # ========================================================
    # 队列轮询 -> UI 更新
    # ========================================================

    def _poll_queue(self):
        """每 100ms 轮询一次消息队列"""
        while not self.msg_queue.empty():
            try:
                msg = self.msg_queue.get_nowait()
                mtype = msg[0]

                if mtype == "log":
                    self._log(msg[1])

                elif mtype == "progress":
                    _, tag, collected, total, page = msg
                    self._upd_tag_progress(tag, collected, total, page)

                elif mtype == "tag_done":
                    _, tag, count, filepath = msg
                    self._on_tag_done(tag, count, filepath)

                elif mtype == "all_done":
                    _, results, files, elapsed = msg
                    self._on_all_done(results, files, elapsed)

                elif mtype == "error":
                    self._log(f"致命错误: {msg[1]}", "error")

                elif mtype == "thread_end":
                    self._on_thread_end()

                # --- 联系方式提取消息 ---
                elif mtype == "contact_log":
                    self._log(msg[1])

                elif mtype == "contact_progress":
                    _, idx, total, name, pid = msg
                    self._upd_contact_progress(idx, total, name, pid)

                elif mtype == "contact_total":
                    self.pb_contact["maximum"] = msg[1]

                elif mtype == "contact_done":
                    _, results, output_file = msg
                    self._on_contact_done(results, output_file)

                elif mtype == "contact_error":
                    self._log(f"❌ 提取错误: {msg[1]}", "error")
                    self.contact_running = False
                    self.btn_contact_start.config(state="normal")
                    self.btn_contact_stop.config(state="disabled")
                    self.btn_contact_skip.config(state="disabled")

                elif mtype == "contact_pause_started":
                    self.btn_contact_skip.config(state="normal")

                elif mtype == "contact_pause_ended":
                    self.btn_contact_skip.config(state="disabled")

            except queue.Empty:
                break
        self.root.after(100, self._poll_queue)

    # ========================================================
    # UI 更新方法
    # ========================================================

    def _log(self, msg: str, tag: str = None):
        """追加日志 (自动着色 + 时间戳 + 自动滚动)"""
        if tag is None:
            # 根据 emoji 自动判断颜色
            tag = self._detect_log_color(msg)

        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{ts}] {msg}\n", tag)
        self.log_text.see("end")

        # 限制日志行数 (保留最后 1500 行)
        lines = int(self.log_text.index("end-1c").split(".")[0])
        if lines > 2000:
            self.log_text.delete("1.0", f"{lines - 1500}.0")

    @staticmethod
    def _detect_log_color(msg: str) -> str:
        """根据消息内容中的 emoji 判断日志颜色"""
        # 成功: 绿色
        for e in ["\u2705", "\U0001f49a", "\U0001f389", "\U0001f680",
                   "\u2713", "\U0001f513"]:
            if e in msg:
                return "success"
        # 警告: 黄色
        for e in ["\u26a0", "\u23f3", "\u23f8", "\u23f1"]:
            if e in msg:
                return "warning"
        # 错误: 红色
        for e in ["\u274c", "\u26d4", "\U0001f510", "\U0001f6d1"]:
            if e in msg:
                return "error"
        # 信息: 蓝色
        for e in ["\U0001f4cc", "\U0001f4e5", "\U0001f4ca", "\U0001f4be",
                   "\U0001f4c4", "\U0001f4c8", "\U0001f4e1", "\U0001f50d",
                   "\U0001f50c"]:
            if e in msg:
                return "info"
        return "normal"

    def _upd_tag_progress(self, tag, collected, total, page):
        """更新当前标签的进度条"""
        self.current_tag = tag
        pct = min(100, int(collected / total * 100)) if total else 0
        self.pb_tag["value"] = pct

        self.lbl_cur.config(
            text=f"当前: {tag} ({self.completed_tags + 1}/{self.total_tags})"
        )
        self.lbl_cur2.config(text=f"{collected}/{total} ({pct}%) | 第{page}页")
        self._upd_stats()

    def _on_tag_done(self, tag, count, filepath):
        """一个标签完成"""
        self.completed_tags += 1
        self.total_collected += count
        if filepath:
            self.files_created.append(filepath)
        self.pb_all["value"] = self.completed_tags
        self.lbl_all.config(text=f"总进度: {self.completed_tags}/{self.total_tags}")
        self.lbl_all2.config(text=f"已采集 {self.total_collected} 人")
        self._upd_stats()

    def _on_all_done(self, results, files, elapsed):
        """全部完成"""
        total = sum(results.values())
        self._log("", "normal")
        self._log("=" * 50, "success")
        self._log("采集任务完成!", "success")
        self._log(f"   总耗时: {self._fmt(elapsed)}", "normal")
        self._log(f"   新增: {len(results)} 标签, {total} 条达人", "normal")
        self._log(f"   文件: {len(files)} 个 Excel", "normal")
        self._log("=" * 50, "success")
        self._set_status("done",
            f"OK 完成 | {total}人 | {len(files)}文件 | {self._fmt(elapsed)}")
        self.pb_tag["value"] = 100

    def _on_thread_end(self):
        """采集线程结束"""
        self.is_running = False
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        current = self.status_var.get()
        if "完成" not in current and "错误" not in current and "停止" not in current:
            self._set_status("ready", "* 就绪")

    def _upd_stats(self):
        """更新统计行"""
        elapsed = int(time.time() - self.start_time) if self.start_time else 0
        files = len(self.files_created)
        self.lbl_stats.config(
            text=f"已采集: {self.total_collected} 人 | "
                 f"文件: {files} 个 | 耗时: {self._fmt(elapsed)}"
        )

    def _tick_timer(self):
        """每秒刷新一次耗时显示"""
        if self.is_running:
            self._upd_stats()
            self.root.after(1000, self._tick_timer)

    def _set_status(self, level, text):
        """更新状态栏"""
        colors = {
            "ready": C_DIM,
            "running": "#00aa00",
            "stopped": "#cc6600",
            "done": "#0066cc",
            "error": "#cc0000",
        }
        self.status_var.set(text)
        self.lbl_status.config(foreground=colors.get(level, C_DIM))

    def _clear_log(self):
        """清空日志"""
        self.log_text.delete("1.0", "end")

    # ========================================================
    # 联系方式提取 UI 更新
    # ========================================================

    def _upd_contact_progress(self, idx, total, name, pid):
        """更新联系方式提取进度"""
        pct = int((idx + 1) / total * 100)
        self.pb_contact["value"] = idx + 1
        self.lbl_contact_progress.config(
            text=f"[{idx+1}/{total}] {name} (ID:{pid}) | "
                 f"已提取{self.contact_extracted}人, 成功{self.contact_success}人"
        )

    def _on_contact_done(self, results, output_file):
        """联系方式提取完成"""
        self.contact_running = False
        total = len(results)
        success = sum(1 for r in results if r.get("phone") or r.get("wechat"))
        elapsed = int(time.time() - self.contact_start_time) if self.contact_start_time else 0

        self._log("", "normal")
        self._log("=" * 50, "success")
        self._log("联系方式提取完成!", "success")
        self._log(f"   总人数: {total} | 提取成功: {success} | "
                  f"耗时: {self._fmt(elapsed)}", "normal")
        self._log(f"   输出: {Path(output_file).name}", "normal")
        self._log("=" * 50, "success")

        self.lbl_contact_progress.config(
            text=f"✅ 完成! {success}/{total} 人成功 | 耗时 {self._fmt(elapsed)}"
        )
        self.pb_contact["value"] = total
        self.btn_contact_start.config(state="normal")
        self.btn_contact_stop.config(state="disabled")
        self.btn_contact_skip.config(state="disabled")
        self.btn_contact_copy.config(state="normal")

    # ========================================================
    # 工具方法
    # ========================================================

    @staticmethod
    def _fmt(secs):
        if secs < 60:
            return f"{secs}秒"
        elif secs < 3600:
            return f"{secs//60}分{secs%60}秒"
        return f"{secs//3600}时{(secs%3600)//60}分"

    def _on_close(self):
        """窗口关闭"""
        if self.is_running:
            if not messagebox.askyesno("确认", "采集正在进行中, 确定退出?"):
                return
            if self.scraper:
                self.scraper.stop()
                time.sleep(1)
        self.root.destroy()

    # ========================================================
    # 启动
    # ========================================================

    def run(self):
        self._log(f"欢迎使用快手达人采集工具（当前产品: {self.current_product.name}）", "info")
        self._log("选择左侧标签后点击 [开始采集]", "normal")
        self._log("首次使用需扫码登录快手, 之后自动保持登录态", "normal")
        self._log("", "normal")
        self.root.mainloop()


# ============================================================
# 产品管理窗口（可视化编辑/新建/删除「人群定义」）
# ============================================================

class ProductManager:
    """弹窗：编辑一份产品的所有筛选与评分参数。

    保存时回调 on_saved(name)；删除时回调 on_saved(剩余第一个产品名)。
    """

    def __init__(self, parent, current_name: str, on_saved):
        self.parent = parent
        self.on_saved = on_saved
        self.original_name = current_name
        prod = ys.load_product(current_name)

        self.win = tk.Toplevel(parent)
        self.win.title(f"产品管理 - {current_name}")
        self.win.geometry("720x640")
        self.win.minsize(640, 560)
        self.win.configure(bg=C_BG)
        self.win.transient(parent)
        self.win.grab_set()

        # 滚动容器
        canvas = tk.Canvas(self.win, bg=C_BG, highlightthickness=0)
        sb = ttk.Scrollbar(self.win, orient="vertical", command=canvas.yview)
        body = ttk.Frame(canvas)
        body.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        f = ttk.Frame(body, padding=10)
        f.pack(fill="both", expand=True)

        self._vars = {}
        self._texts = {}

        def add_entry(row, label, key, value="", w=40):
            ttk.Label(f, text=label).grid(row=row, column=0, sticky="w", pady=2)
            v = tk.StringVar(value=value)
            self._vars[key] = v
            ttk.Entry(f, textvariable=v, width=w).grid(row=row, column=1, sticky="we", pady=2)

        def add_combo(row, label, key, values, value=""):
            ttk.Label(f, text=label).grid(row=row, column=0, sticky="w", pady=2)
            v = tk.StringVar(value=value)
            self._vars[key] = v
            ttk.Combobox(f, textvariable=v, values=values, state="readonly", width=14)\
                .grid(row=row, column=1, sticky="w", pady=2)

        def add_check(row, label, key, value=False):
            ttk.Label(f, text=label).grid(row=row, column=0, sticky="w", pady=2)
            v = tk.BooleanVar(value=value)
            self._vars[key] = v
            ttk.Checkbutton(f, variable=v).grid(row=row, column=1, sticky="w", pady=2)

        def add_text(row, label, key, value="", h=4):
            ttk.Label(f, text=label).grid(row=row, column=0, sticky="nw", pady=2)
            t = tk.Text(f, width=70, height=h, wrap="word")
            t.insert("1.0", value)
            t.grid(row=row, column=1, sticky="we", pady=2)
            self._texts[key] = t

        r = 0
        ttk.Label(f, text="基础信息", font=F_TITLE).grid(row=r, column=0, columnspan=2, sticky="w")
        r += 1
        add_entry(r, "产品名称", "name", prod.name); r += 1
        add_entry(r, "关联度列名", "assoc_field", prod.assoc_field); r += 1
        add_entry(r, "受众列名", "audience_label", prod.audience_label); r += 1
        add_combo(r, "目标性别", "target_gender", ["女", "男", "无"], prod.target_gender); r += 1
        add_entry(r, "近30日分销销售额阈值(0=不启用)", "min_dist_sales_30d",
                  str(prod.min_dist_sales_30d), w=14); r += 1
        add_check(r, "无销售额数据也剔除", "drop_no_sales_data", prod.drop_no_sales_data); r += 1
        add_check(r, "采集粉丝画像(取消则跳过 per-达人查询，更快更省接口)",
                  "need_portrait", prod.need_portrait); r += 1
        add_entry(r, "推荐标签(逗号分隔)", "recommended_tags",
                  "，".join(prod.recommended_tags or [])); r += 1

        r += 1
        ttk.Label(f, text="三级关联关键词（每行一个）", font=F_TITLE)\
            .grid(row=r, column=0, columnspan=2, sticky="w"); r += 1
        ttk.Label(f, text="提示：三级关键词全部留空 = 通用手动模式（不做关联剔除，仅按销售额门槛 + 你手动选的标签筛选）",
                   foreground="#888888", wraplength=560)\
            .grid(row=r, column=0, columnspan=2, sticky="w"); r += 1
        add_text(r, "强关联", "kw_tier1", "\n".join(prod.kw_tier1)); r += 1
        add_text(r, "中关联", "kw_tier2", "\n".join(prod.kw_tier2)); r += 1
        add_text(r, "弱关联", "kw_tier3", "\n".join(prod.kw_tier3), h=3); r += 1

        r += 1
        ttk.Label(f, text="高级：品类/标签契合分（JSON，谨慎修改）", font=F_TITLE)\
            .grid(row=r, column=0, columnspan=2, sticky="w"); r += 1
        add_text(r, "带货品类契合分", "channel_fit",
                 json.dumps(prod.channel_fit, ensure_ascii=False, indent=0), h=4); r += 1
        add_text(r, "带货品类映射", "channel_map",
                 json.dumps(prod.channel_map, ensure_ascii=False, indent=0), h=3); r += 1
        add_text(r, "内容标签契合分", "tag_fit_score",
                 json.dumps(prod.tag_fit_score, ensure_ascii=False, indent=0), h=4); r += 1

        # 按钮行
        btn = ttk.Frame(f)
        btn.grid(row=r, column=0, columnspan=2, pady=10)
        ttk.Button(btn, text="保存", command=self._on_save).pack(side="left", padx=6)
        ttk.Button(btn, text="另存为新品", command=self._on_save_as_new).pack(side="left", padx=6)
        ttk.Button(btn, text="删除该产品", command=self._on_delete).pack(side="left", padx=6)
        ttk.Button(btn, text="取消", command=self.win.destroy).pack(side="left", padx=6)

        f.columnconfigure(1, weight=1)
        # 非阻塞模态：grab_set 已锁定主窗口，事件由主 mainloop 驱动；
        # 用户点 取消/保存/删除 时本窗口自行 destroy。

    # ---- 辅助 ----
    @staticmethod
    def _split_lines(text):
        return [x.strip() for x in text.strip().splitlines() if x.strip()]

    def _parse_text(self, key, default):
        t = self._texts.get(key)
        if not t:
            return default
        raw = t.get("1.0", "end").strip()
        if not raw:
            return default
        try:
            return json.loads(raw)
        except Exception as e:
            raise ValueError(f"{key} 不是合法 JSON: {e}")

    def _collect(self, name_override=None):
        name = (name_override or self._vars["name"].get()).strip()
        if not name:
            raise ValueError("产品名称不能为空")
        rec = [x.strip() for x in self._vars["recommended_tags"].get().replace("，", ",").split(",") if x.strip()]
        try:
            sales = int(self._vars["min_dist_sales_30d"].get() or 0)
        except ValueError:
            raise ValueError("分销销售额阈值必须是整数")
        return ys.ProductConfig(
            name=name,
            assoc_field=self._vars["assoc_field"].get().strip() or "产品关联",
            audience_label=self._vars["audience_label"].get().strip() or "目标受众",
            recommended_tags=rec,
            tag_fit_score=self._parse_text("tag_fit_score", {}),
            channel_fit=self._parse_text("channel_fit", {}),
            channel_map=self._parse_text("channel_map", {}),
            kw_tier1=self._split_lines(self._texts["kw_tier1"].get("1.0", "end")),
            kw_tier2=self._split_lines(self._texts["kw_tier2"].get("1.0", "end")),
            kw_tier3=self._split_lines(self._texts["kw_tier3"].get("1.0", "end")),
            target_gender=self._vars["target_gender"].get() or "女",
            min_dist_sales_30d=sales,
            drop_no_sales_data=self._vars["drop_no_sales_data"].get(),
            need_portrait=self._vars["need_portrait"].get(),
        )

    def _on_save(self):
        try:
            cfg = self._collect()
        except ValueError as e:
            messagebox.showerror("校验失败", str(e))
            return
        # 改名时删除旧文件，避免残留
        if cfg.name != self.original_name:
            old_fp = ys.PRODUCTS_DIR / f"{self.original_name}.json"
            if old_fp.exists():
                old_fp.unlink()
        ys.save_product(cfg)
        messagebox.showinfo("已保存", f"产品「{cfg.name}」已保存")
        self.win.destroy()
        self.on_saved(cfg.name)

    def _on_save_as_new(self):
        try:
            cfg = self._collect(name_override=f"{self._vars['name'].get().strip()}_副本")
        except ValueError as e:
            messagebox.showerror("校验失败", str(e))
            return
        ys.save_product(cfg)
        messagebox.showinfo("已保存", f"已另存为新品「{cfg.name}」")
        self.win.destroy()
        self.on_saved(cfg.name)

    def _on_delete(self):
        name = self._vars["name"].get().strip()
        if name == self.original_name and messagebox.askyesno("确认删除",
                f"确定删除产品「{name}」？此操作不可恢复。"):
            fp = ys.PRODUCTS_DIR / f"{name}.json"
            if fp.exists():
                fp.unlink()
            names = ys.list_products()
            fallback = names[0] if names else "颜阿娇"
            messagebox.showinfo("已删除", f"产品「{name}」已删除")
            self.win.destroy()
            self.on_saved(fallback)


# ============================================================
# 入口
# ============================================================

if __name__ == "__main__":
    try:
        app = ScraperGUI()
        app.run()
    except Exception as e:
        # 任何启动期异常都不要静默闪退：弹窗 + 写日志，方便定位
        import traceback
        tb = traceback.format_exc()
        try:
            crash_log = Path(__file__).parent / "gui_crash.log"
            crash_log.write_text(
                f"时间: {datetime.now():%Y-%m-%d %H:%M:%S}\n\n{tb}",
                encoding="utf-8",
            )
        except Exception:
            crash_log = None
        try:
            _rk = tk.Tk()
            _rk.withdraw()
            msg = f"GUI 启动失败：\n\n{tb}"
            if crash_log:
                msg += f"\n\n完整错误已保存到：\n{crash_log}"
            messagebox.showerror("启动失败", msg)
            _rk.destroy()
        except Exception:
            # tkinter 也起不来时，退回控制台输出
            sys.stderr.write(tb + "\n")
        try:
            if crash_log:
                _show_err("启动失败", tb)  # 原生弹窗，必定可见
                os.startfile(str(crash_log))
        except Exception:
            pass
        raise
