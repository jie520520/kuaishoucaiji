# -*- coding: utf-8 -*-
"""
颜阿娇 - 快手达人采集工具 v11.7（多产品配置化 + 页面级先筛后拉 + 分销销售额过滤 + CDP接管）
===============================
独立封装版本：一个脚本搞定全部达人采集。
换产品只需新增/修改 products/<产品名>.json（人群定义），无需改代码。

v11.5 升级：平台级先筛后拉 — 在达人广场页面真实点击筛选条件再拉数据
  - 采集前在页面上设置：达人地域/达人性别/粉丝年龄/粉丝性别/粉丝城市划分/有联系方式
  - 拦截页面自动发出的promoter/list请求体，复用快手自己构造的完整筛选参数翻页
  - 平台直接返回筛选后的达人（如：只返回粉丝年龄31-40岁的女性达人）
  - 不再需要"拉697人→查697人详情→本地筛到60人"的过度查询
  - 页面筛选失败时自动降级为 v11.4 本地预筛选，不影响功能

v11.4 升级：预筛选优化 — 在查询详情前先用列表数据过滤
  - 性别预筛：列表API已含gender字段，查详情前直接过滤（省50%详情查询）
  - 减肥关联预筛：从商品数据匹配关键词，无关联的不查详情（省90%详情查询）
  - API级性别筛选：尝试在promoter/list接口传gender参数，不支持则自动降级
  - 效果：697人 → 预筛选60人 → 只查60人详情（原来查697人，省6分钟）

功能：
  1. 按内容标签精准筛选达人（37个标签全覆盖）
  2. 自动产品契合度评分（减肥关联 + 品类 + 女性受众 + 粉丝性价比）
  3. 评分排序 → 条件格式Excel输出
  4. 登录态持久化（一次扫码，长期有效）
  5. 断点续传（已采集标签自动跳过）
  6. 限流自动重试

使用方式：
  python yanajiao_scraper.py --all        # 采集全部37个标签
  python yanajiao_scraper.py --rec        # 采集推荐标签（健康/运动/美妆等）
  python yanajiao_scraper.py --tag 健康    # 采集指定标签（逗号分隔多个）
  python yanajiao_scraper.py              # 交互模式
"""

import os
import sys
import io
import time
import json
import argparse
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    try:
        if hasattr(sys.stdout, 'buffer') and not isinstance(sys.stdout, io.TextIOWrapper):
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        if hasattr(sys.stderr, 'buffer') and not isinstance(sys.stderr, io.TextIOWrapper):
            sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
        # 关键：强制行缓冲，让输出立即显示到黑窗口
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(line_buffering=True)
    except Exception:
        pass  # 模块导入或管道模式时忽略

# 全局时间标记
_START_TIME = time.time()


def _elapsed() -> str:
    """返回从脚本启动到现在的耗时"""
    secs = int(time.time() - _START_TIME)
    if secs < 60:
        return f"{secs}秒"
    elif secs < 3600:
        return f"{secs//60}分{secs%60}秒"
    return f"{secs//3600}时{(secs%3600)//60}分"


def _ts() -> str:
    """当前时间戳，用于日志前缀"""
    return datetime.now().strftime("%H:%M:%S")


def log(msg: str, end: str = "\n"):
    """带时间戳 + 即刻刷新的打印"""
    text = f"[{_ts()}] {msg}"
    print(text, end=end, flush=True)

from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ============================================================
# 配置常量
# ============================================================

# ============================================================
# 产品配置（多产品支持：每个产品一份「人群定义」）
# ============================================================
# 设计目标：产品会持续更新（颜阿娇只是第一个），以后换产品（母婴/美妆/食品…）
# 只需新增/修改一份 JSON 配置，无需改动代码。换产品 = 换一份配置。
import json
from dataclasses import dataclass, field, asdict

# 产品配置目录（与脚本同目录下的 products/）
PRODUCTS_DIR = Path(__file__).parent / "products"


@dataclass
class ProductConfig:
    """一份产品的「人群定义」：决定找什么样的达人、怎么评分、阈值多少。

    字段说明：
      name             产品名（用于输出文件名前缀、日志、下拉显示）
      assoc_field      表格中「关联度」列名（如 减肥关联 / 母婴关联）
      audience_label   表格中受众列名（如 女性受众）
      recommended_tags 推荐优先采集的内容标签
      tag_fit_score    内容标签 -> 契合基础分
      channel_fit      带货品类 -> 契合分
      channel_map      带货品类 -> 内容标签
      kw_tier1/2/3     强/中/弱 关联关键词（从达人热卖商品名匹配）
      tier_scores      各档 {每词分, 上限}
      female_nick_chars/male_nick_chars  目标性别的昵称特征字
      target_gender    目标受众性别：女 / 男 / 无
      min_dist_sales_30d  近30日分销销售额(场均)阈值，0/None=不启用
      drop_no_sales_data  无销售额数据是否也剔除
      score_weights    四个评分维度满分权重
    """
    name: str = "颜阿娇"
    assoc_field: str = "产品关联"
    audience_label: str = "女性受众"
    recommended_tags: list = field(default_factory=list)
    tag_fit_score: dict = field(default_factory=dict)
    channel_fit: dict = field(default_factory=dict)
    channel_map: dict = field(default_factory=dict)
    kw_tier1: list = field(default_factory=list)
    kw_tier2: list = field(default_factory=list)
    kw_tier3: list = field(default_factory=list)
    tier_scores: dict = field(default_factory=lambda: {1: [15, 45], 2: [10, 20], 3: [5, 10]})
    female_nick_chars: list = field(default_factory=list)
    male_nick_chars: list = field(default_factory=list)
    target_gender: str = "女"
    min_dist_sales_30d: int = 0
    drop_no_sales_data: bool = False
    need_portrait: bool = True  # False=跳过 per-达人画像查询（通用极简模式，省 N×2 次接口）
    score_weights: dict = field(default_factory=lambda: {"assoc": 50, "channel": 20, "gender": 15, "fans": 15})

    @staticmethod
    def _norm_tier(t):
        # JSON 的 key 是字符串，统一转回 int
        if isinstance(t, dict):
            return {int(k): v for k, v in t.items()}
        return t

    @classmethod
    def from_dict(cls, d: dict) -> "ProductConfig":
        d = dict(d)
        if "tier_scores" in d and isinstance(d["tier_scores"], dict):
            d["tier_scores"] = cls._norm_tier(d["tier_scores"])
        return cls(**d)

    def to_dict(self) -> dict:
        return asdict(self)

    @property
    def has_assoc_rules(self) -> bool:
        """是否配置了「产品关联」匹配规则（关键词 / 品类契合分）。

        未配置 = 通用手动模式：不做关联剔除，达人只按「销售额门槛 + 手动选标签」筛选，
        关联列填 "-" 表示不参与。配置了任意一项则恢复「无关联即剔除」行为。
        """
        return bool(self.kw_tier1 or self.kw_tier2 or self.kw_tier3 or self.channel_fit)


# 内置默认产品（首次运行写入 products/ 目录；之后以目录内 JSON 为准）
_DEFAULT_PRODUCTS = {
    "颜阿娇": {
        "name": "颜阿娇",
        "assoc_field": "减肥关联",
        "audience_label": "女性受众",
        "recommended_tags": ["健康", "运动", "美妆", "穿搭", "亲子", "生活", "颜值", "舞蹈"],
        "tag_fit_score": {"健康":20,"运动":12,"美妆":10,"穿搭":7,"亲子":6,"生活":6,"颜值":5,"舞蹈":3,"情感":2,"美食":2,"旅游":1,"搞笑":1,"才艺":1,"高新数码":0,"读书":0},
        "channel_fit": {"营养健康":20,"养生保健":20,"运动户外":12,"美妆护肤":10,"个护清洁":10,"母婴玩具":10,"女装女鞋":5,"男装男鞋":3,"家居百货":3,"零食饮料":2,"生鲜食品":2,"茶叶酒水":2},
        "channel_map": {"营养健康":"健康","养生保健":"健康","运动户外":"运动","美妆护肤":"美妆","个护清洁":"生活","女装女鞋":"穿搭","男装男鞋":"穿搭","母婴玩具":"亲子","家居百货":"生活","零食饮料":"美食","生鲜食品":"美食","茶叶酒水":"美食","珠宝奢品":"穿搭","数码家电":"高新数码","图书文娱":"读书"},
        "kw_tier1": ["减肥","瘦身","减脂","燃脂","减重","掉秤","塑形","纤体","轻体","瘦腿","瘦肚子","瘦腰","瘦脸","瘦臂","体态管理","体重管理","控体","窈窕","窈轻","减肥咖啡","瘦身咖啡","燃脂咖啡","黑咖啡减肥","左旋肉碱","白芸豆","淀粉酶","糖阻断","碳水阻断","代餐","饱腹","低卡代餐","轻食代餐","酵素","果蔬酵素","复合酵素","排油","刮油","燃卡","瘦素"],
        "kw_tier2": ["黑咖啡","冻干咖啡","速溶咖啡","益生菌","膳食纤维","菊粉","低聚糖","控糖","低卡","低糖","代糖","0卡","零卡","祛湿","健脾","利水","消脂","代谢","基础代谢","新陈代谢","草本","植物萃取","天然植物","膳食","营养代餐","蛋白粉","润肠","排毒","抗糖","抗氧化","胶原蛋白肽"],
        "kw_tier3": ["健康","养生","保健","营养","美容","养颜","护肤","变美","身材","体型","曲线","运动","健身","瑜伽","产后","恢复","辣妈"],
        "tier_scores": {"1":[15,45],"2":[10,20],"3":[5,10]},
        "female_nick_chars": ["妈","姐","妹","娘","女","美","花","丽","娜","芳","娟","婷","静","敏","雪","琳","玲","菲","萱","琪"],
        "male_nick_chars": ["哥","弟","男","伟","强","杰","磊","勇","军","斌","峰","涛","龙","虎","刚"],
        "target_gender": "女",
        "min_dist_sales_30d": 5000,
        "drop_no_sales_data": False,
        "need_portrait": True,
        "score_weights": {"assoc":50,"channel":20,"gender":15,"fans":15},
    },
    "示例_母婴": {
        "name": "示例_母婴",
        "assoc_field": "母婴关联",
        "audience_label": "女性受众",
        "recommended_tags": ["亲子","生活","美妆","穿搭","美食","健康","颜值","舞蹈"],
        "tag_fit_score": {"亲子":20,"生活":12,"美妆":10,"穿搭":7,"美食":6,"健康":6,"颜值":5,"舞蹈":3,"情感":2,"运动":2,"旅游":1,"搞笑":1,"才艺":1,"高新数码":0,"读书":0},
        "channel_fit": {"母婴玩具":20,"个护清洁":12,"美妆护肤":10,"女装女鞋":8,"家居百货":8,"营养健康":6,"生鲜食品":5,"零食饮料":5,"运动户外":3,"图书文娱":3},
        "channel_map": {"母婴玩具":"亲子","个护清洁":"生活","美妆护肤":"美妆","女装女鞋":"穿搭","家居百货":"生活","营养健康":"健康","生鲜食品":"美食","零食饮料":"美食","运动户外":"运动","图书文娱":"读书"},
        "kw_tier1": ["婴儿","宝宝","母婴","孕婴","孕妇","新生儿","纸尿裤","奶瓶","辅食","早教","童装","孕产","待产","月子","哄睡","断奶","孕吐","胎动"],
        "kw_tier2": ["儿童","育儿","带娃","亲子","辣妈","宝妈","产后","恢复","益智","玩具","绘本","奶粉","保湿","护肤","安抚","学步"],
        "kw_tier3": ["家庭","生活","健康","营养","变美","身材","恢复","亲子"],
        "tier_scores": {"1":[15,45],"2":[10,20],"3":[5,10]},
        "female_nick_chars": ["妈","姐","妹","娘","女","美","花","丽","娜","芳","娟","婷","静","敏","雪","琳","玲","菲","萱","琪"],
        "male_nick_chars": ["哥","弟","男","伟","强","杰","磊","勇","军","斌","峰","涛","龙","虎","刚"],
        "target_gender": "女",
        "min_dist_sales_30d": 3000,
        "drop_no_sales_data": False,
        "need_portrait": True,
        "score_weights": {"assoc":50,"channel":20,"gender":15,"fans":15},
    },
    # 通用手动模式模板：只设销售额门槛，其余全留空
    # 勾选标签由用户在 GUI 手动选择，不做关键词/品类关联剔除
    "通用_仅销售额": {
        "name": "通用_仅销售额",
        "assoc_field": "产品关联",
        "audience_label": "目标受众",
        "recommended_tags": [],
        "tag_fit_score": {},
        "channel_fit": {},
        "channel_map": {},
        "kw_tier1": [],
        "kw_tier2": [],
        "kw_tier3": [],
        "tier_scores": {"1":[15,45],"2":[10,20],"3":[5,10]},
        "female_nick_chars": [],
        "male_nick_chars": [],
        "target_gender": "无",
        "min_dist_sales_30d": 5000,
        "drop_no_sales_data": False,
        "need_portrait": False,
        "score_weights": {"assoc":50,"channel":20,"gender":15,"fans":15},
    },
}


def ensure_default_products():
    """首次运行：把内置默认产品写入 products/ 目录（已存在则不覆盖）"""
    PRODUCTS_DIR.mkdir(parents=True, exist_ok=True)
    for name, cfg in _DEFAULT_PRODUCTS.items():
        fp = PRODUCTS_DIR / f"{name}.json"
        if not fp.exists():
            fp.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")


def list_products() -> list:
    """返回可用产品名列表（目录内 JSON + 内置默认，去重，保序；主产品颜阿娇置顶）"""
    ensure_default_products()
    names = []
    for fp in sorted(PRODUCTS_DIR.glob("*.json")):
        try:
            d = json.loads(fp.read_text(encoding="utf-8"))
            n = d.get("name") or fp.stem
            if n not in names:
                names.append(n)
        except Exception:
            pass
    for n in _DEFAULT_PRODUCTS:
        if n not in names:
            names.append(n)
    # 主产品颜阿娇优先置顶
    if "颜阿娇" in names:
        names.remove("颜阿娇")
        names.insert(0, "颜阿娇")
    return names


def load_product(name: str | None = None) -> "ProductConfig":
    """载入指定产品配置；name 为空或找不到时回退到主产品（颜阿娇）/ 第一个可用产品。

    备注：颜阿娇是首个上线产品，作为默认主产品优先。
    """
    ensure_default_products()
    if name:
        fp = PRODUCTS_DIR / f"{name}.json"
        if fp.exists():
            try:
                return ProductConfig.from_dict(json.loads(fp.read_text(encoding="utf-8")))
            except Exception:
                pass
    # 回退：优先主产品 颜阿娇，其次目录第一个 / 内置默认第一个
    names = list_products()
    if not names:
        return ProductConfig.from_dict(_DEFAULT_PRODUCTS["颜阿娇"])
    preferred = "颜阿娇" if "颜阿娇" in names else names[0]
    fp = PRODUCTS_DIR / f"{preferred}.json"
    if fp.exists():
        try:
            return ProductConfig.from_dict(json.loads(fp.read_text(encoding="utf-8")))
        except Exception:
            pass
    return ProductConfig.from_dict(_DEFAULT_PRODUCTS.get(preferred, _DEFAULT_PRODUCTS["颜阿娇"]))


def save_product(cfg: "ProductConfig"):
    """保存（新增/覆盖）产品配置到 products/<name>.json"""
    PRODUCTS_DIR.mkdir(parents=True, exist_ok=True)
    fp = PRODUCTS_DIR / f"{cfg.name}.json"
    fp.write_text(json.dumps(cfg.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")


def delete_product(name: str) -> bool:
    """删除产品配置（内置默认不可删，防止误删）"""
    if name in _DEFAULT_PRODUCTS:
        return False
    fp = PRODUCTS_DIR / f"{name}.json"
    if fp.exists():
        fp.unlink()
        return True
    return False


class Config:
    """全局配置"""
    DAREN_URL = "https://cps.kwaixiaodian.com/zone/daren-match/daren-square-pro"
    API_URL = "https://cps.kwaixiaodian.com/distribute/pc/seller/promoter/list"
    API_DETAIL_URL = "https://cps.kwaixiaodian.com/distribute/pc/seller/promoter/info"
    API_FANS_URL = "https://cps.kwaixiaodian.com/distribute/pc/seller/promoter/fans/info"
    BASE_DIR = Path(__file__).parent
    USER_DATA = BASE_DIR / ".kuaishou_browser_data"
    OUTPUT_DIR = BASE_DIR / "采集结果"
    PAGE_SIZE = 20

    # 全部37个标签（平台维度，固定不变）
    ALL_TAGS = [
        "三农", "二次元", "亲子", "随手拍", "生活", "穿搭", "美妆", "美食",
        "旅游", "健康", "游戏", "情感", "资讯", "颜值", "运动", "高新数码",
        "动物", "汽车", "音乐", "影视和短剧", "法律", "才艺", "明星娱乐",
        "军事", "教育", "宗教", "读书", "房产家居", "摄影", "舞蹈",
        "搞笑", "财经", "星座命理", "奇人异象", "科学", "历史", "其他",
    ]

    # 标签 → API数字ID（平台维度，固定不变）
    TAG_IDS = {
        "随手拍": [1], "游戏": [2], "二次元": [3], "情感": [4],
        "资讯": [5], "穿搭": [6], "运动": [7], "汽车": [8],
        "音乐": [9], "颜值": [10], "影视和短剧": [11], "亲子": [12],
        "法律": [13], "生活": [14], "才艺": [15], "明星娱乐": [16],
        "军事": [17], "教育": [18], "三农": [19], "宗教": [20],
        "读书": [21], "房产家居": [22], "美妆": [23], "健康": [24],
        "摄影": [25], "动物": [26], "舞蹈": [27], "搞笑": [28],
        "美食": [29], "财经": [30], "星座命理": [31], "高新数码": [32],
        "奇人异象": [33], "旅游": [34], "科学": [35], "历史": [36],
        "其他": [36, 37],
    }

    # ── 产品相关配置（推荐标签 / 关联关键词 / 品类分 / 目标性别 / 销售额阈值等）──
    # 已外置到 ProductConfig，存放在 products/<产品名>.json。
    # 换产品只需新增/修改对应 JSON（见 load_product / save_product），无需改代码。


# ============================================================
# 评分引擎（v11: 减肥赛道聚焦）
# ============================================================

class Scorer:
    """产品契合度评分（0-100）

    评分体系（全部来自产品配置 ProductConfig）：
      1. 产品关联度（核心，三级关键词匹配达人热卖商品）
      2. 带货品类关联度
      3. 目标受众匹配（受产品 target_gender 控制）
      4. 粉丝性价比（平台维度）
    """

    @staticmethod
    def score(item: dict, tag: str, product: "ProductConfig") -> tuple:
        """
        返回 (总分, 评分明细, 匹配关键词列表)
        评分完全来自产品配置 product（关联关键词 / 品类分 / 目标性别 / 权重）。
        """
        parts = []
        total = 0
        matched_kws = []

        # ===== 1. 产品关联度（核心）— 从达人热卖商品名匹配三级关键词 =====
        items_text = str(item.get("top_items_text", ""))
        wl_score = 0
        ts = product.tier_scores
        per1, cap1 = ts.get(1, [15, 45])
        per2, cap2 = ts.get(2, [10, 20])
        per3, cap3 = ts.get(3, [5, 10])

        # 强关联
        t1 = [k for k in product.kw_tier1 if k in items_text]
        wl_score += min(len(t1) * per1, cap1)
        matched_kws.extend(t1[:4])

        # 中关联
        t2 = [k for k in product.kw_tier2 if k in items_text]
        t2_score = min(len(t2) * per2, cap2)
        wl_score += t2_score
        matched_kws.extend(t2[:(6 - len(matched_kws))])

        # 弱关联
        t3 = [k for k in product.kw_tier3 if k in items_text]
        t3_score = min(len(t3) * per3, cap3)
        wl_score += t3_score
        needed = 6 - len(matched_kws)
        if needed > 0:
            matched_kws.extend(t3[:needed])

        total += wl_score
        detail_wl = []
        if t1:
            detail_wl.append(f"强:{'/'.join(t1[:2])}({min(len(t1)*per1,cap1)})")
        if t2:
            detail_wl.append(f"中:{'/'.join(t2[:2])}({t2_score})")
        if t3:
            detail_wl.append(f"弱:{'/'.join(t3[:2])}({t3_score})")
        if not detail_wl:
            detail_wl.append(f"{product.assoc_field}:无(0)")
        parts.append("|".join(detail_wl))

        # ===== 2. 带货品类关联度 =====
        ch = str(item.get("带货品类", ""))
        cat_score = product.channel_fit.get(ch, 0)
        total += cat_score
        parts.append(f"品类:{cat_score}")

        # ===== 3. 目标受众匹配（受产品 target_gender 控制） =====
        gender = str(item.get("达人性别", ""))
        nickname = str(item.get("昵称", ""))
        if product.target_gender == "无":
            fs = 8  # 不偏好特定性别，给中性分
        elif gender == product.target_gender:
            fs = 15
        elif gender:
            fs = 3  # 已知性别但不匹配目标
        else:
            # 性别未知，按昵称特征字推断
            if product.target_gender == "女" and any(c in nickname for c in product.female_nick_chars):
                fs = 8
            elif product.target_gender == "男" and any(c in nickname for c in product.male_nick_chars):
                fs = 8
            else:
                fs = 3
        total += fs
        parts.append(f"受众:{fs}")

        # ===== 4. 粉丝性价比（平台维度，不依赖产品） =====
        fans = item.get("粉丝数_raw", 0)
        if 100000 <= fans <= 5000000:
            fsc = 15
        elif 50000 <= fans < 100000:
            fsc = 10
        elif 5000000 < fans <= 10000000:
            fsc = 8
        elif fans > 0:
            fsc = 3
        else:
            fsc = 0
        total += fsc
        parts.append(f"粉丝:{fsc}")

        return total, " | ".join(parts), matched_kws[:6]

    @staticmethod
    def level(score: int, product: "ProductConfig") -> str:
        """等级文案（带产品名，通用化）"""
        if score >= 75:
            return f"★★★★★ {product.name}核心"
        if score >= 60:
            return "★★★★ 高度契合"
        if score >= 45:
            return "★★★ 推荐合作"
        if score >= 30:
            return "★★ 可考虑"
        if score >= 15:
            return "★ 弱关联"
        return "不推荐"


# ============================================================
# 核心采集器
# ============================================================

class YanajiaoScraper:
    """
    快手达人采集器（多产品配置化）
    ── 按内容标签精准筛选 + 产品契合度评分(人群定义) + 条件格式Excel输出
    构造时传入产品名即切换人群（默认载入 products/ 第一个产品）。
    """

    def __init__(self, product_name: str | None = None):
        # 载入产品配置（换产品 = 换一份配置）；默认载入第一个可用产品
        self.product = load_product(product_name)
        self.playwright = None
        self.context = None
        self.page = None
        self._collector_name = (f"YanajiaoScraper v11.7（多产品配置化「{self.product.name}」"
                                 f" + 页面级先筛后拉 + 分销销售额过滤）")
        self._tag_start_time = 0  # 当前标签开始时间
        # GUI 回调钩子（不设则用模块级 log 函数）
        self._stop = False
        self._on_progress = None       # callback(tag, collected, total, page)
        self._on_tag_done = None       # callback(tag, count, filepath)
        self._on_all_done = None       # callback(results, files, elapsed)
        # 可配置延时（GUI 可调）
        self.page_delay = 2.0          # 页间延迟（秒）
        self.tag_delay = 5             # 标签间延迟（秒）
        self.page_size = Config.PAGE_SIZE  # 每页拉取数量（默认20，可试50）
        # 达人类型筛选（0=全部, 1=直播达人, 2=视频达人）
        self.promoter_type = 0
        # 是否仅采集有联系方式的达人
        self.has_contact = True
        # 采集后筛选条件（由 GUI 传入）
        self.filter_config = {}
        # API级性别筛选状态（fetch_tag中设置）
        self._api_gender_tried = False
        self._api_gender_failed = False
        # 页面级筛选状态（v11.5: 在页面上真实点击筛选后拦截到的API参数）
        self._page_filter_params = None      # 拦截到的promoter/list请求体参数
        self._page_filter_active = False     # 页面级筛选是否生效

    def stop(self):
        """请求停止采集（在当前页完成后生效）"""
        self._stop = True
        log("🛑 收到停止信号，将在当前页完成后停止...")

    # ---------- 浏览器管理 ----------

    def start_browser(self) -> bool:
        """启动浏览器（非 headless，方便扫码登录）。

        设置环境变量 KS_CDP=1 时改为接管已运行的真实 Chrome（需以
        --remote-debugging-port=9222 启动），规避 Playwright 自带 Chromium
        被快手风控踢出登录的问题。失败自动回退到本地 Chromium。
        """
        log(f"🚀 启动浏览器...")
        self._cdp_mode = False
        if os.environ.get("KS_CDP") == "1":
            try:
                return self._start_browser_cdp()
            except Exception as e:
                log(f"⚠️ CDP 接管失败，回退到本地 Chromium: {e}")
        Config.USER_DATA.mkdir(parents=True, exist_ok=True)

        self.playwright = sync_playwright().start()
        self.context = self.playwright.chromium.launch_persistent_context(
            user_data_dir=str(Config.USER_DATA),
            headless=False,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
            viewport={"width": 1440, "height": 900},
            locale="zh-CN",
        )
        self.page = self.context.new_page()
        log("✅ 浏览器已启动")
        return True

    def _start_browser_cdp(self) -> bool:
        """v11.6: 接管已运行的真实 Chrome（CDP）。登录态与用户日常浏览完全一致。"""
        port = os.environ.get("KS_CDP_PORT", "9222")
        endpoint = f"http://127.0.0.1:{port}"
        log(f"🔗 接管真实 Chrome (CDP {endpoint})...")
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.connect_over_cdp(endpoint)
        self.context = self.browser.contexts[0] if self.browser.contexts else self.browser.new_context()
        self.page = self.context.pages[0] if self.context.pages else self.context.new_page()
        if "kwaixiaodian" not in self.page.url:
            self.page.goto(Config.DAREN_URL, wait_until="domcontentloaded", timeout=30000)
            time.sleep(3)
        self._cdp_mode = True
        log("✅ CDP 接管成功（真实 Chrome，登录态保留）")
        return True

    def navigate_to_daren_square(self) -> bool:
        """导航到达人广场页面"""
        log("📡 导航至达人广场...")
        try:
            self.page.goto(Config.DAREN_URL, timeout=60000, wait_until="domcontentloaded")
            time.sleep(3)
            log("✅ 页面加载完成")
            return True
        except Exception as e:
            log(f"⚠️ 导航超时但继续: {e}")
            return True

    def wait_for_login(self, timeout_seconds: int = 180) -> bool:
        """等待用户扫码登录"""
        # 先检查当前是否已登录
        log("🔍 检查登录状态...")
        try:
            body = self.page.locator("body").inner_text(timeout=3000)
            url_lower = self.page.url.lower()
            if "达人广场" in body and "login" not in url_lower:
                log("💚 已有有效登录态，无需扫码")
                time.sleep(2)
                return True
            if "快手" in body and "login" not in url_lower:
                log("🔓 已登录，跳转至达人广场...")
                self.page.goto(Config.DAREN_URL, timeout=30000, wait_until="domcontentloaded")
                time.sleep(3)
                body2 = self.page.locator("body").inner_text(timeout=3000)
                if "达人广场" in body2:
                    log("✅ 已到达达人广场")
                    time.sleep(2)
                    return True
        except Exception:
            pass

        log("")
        log("=" * 50)
        log("  📱 请在浏览器中完成快手扫码登录")
        log("  登录成功后脚本将自动继续")
        log("=" * 50)

        for i in range(timeout_seconds // 2):
            time.sleep(2)
            try:
                url_lower = self.page.url.lower()
                # 已离开登录页 = 登录成功
                if "login" not in url_lower:
                    log("🔓 检测到登录完成，导航至达人广场...")
                    time.sleep(2)
                    self.page.goto(Config.DAREN_URL, timeout=30000, wait_until="domcontentloaded")
                    time.sleep(3)
                    try:
                        body = self.page.locator("body").inner_text(timeout=3000)
                        if "达人广场" in body or "筛选" in body:
                            log("🎉 登录成功！开始采集...")
                            time.sleep(3)
                            return True
                    except Exception:
                        pass
                    if "daren-square" in url_lower or "daren-match" in self.page.url:
                        log("🎉 登录成功！开始采集...")
                        time.sleep(3)
                        return True
            except Exception:
                pass
            if i % 15 == 14:
                log(f"  ⏳ 等待扫码中... (已等 {(i+1)*2} 秒)")

        log("❌ 登录超时，未完成登录")
        return False

    def relogin(self) -> bool:
        """重新登录"""
        log("🔄 需要重新登录，请在浏览器中扫码...")
        self.page.goto(Config.DAREN_URL, timeout=60000, wait_until="domcontentloaded")
        return self.wait_for_login(timeout_seconds=120)

    def close(self):
        """关闭浏览器"""
        log("🔒 关闭浏览器...")
        # CDP 模式接管的是用户真实 Chrome，绝不关闭其上下文（否则会关掉用户浏览器）
        if self.context and not getattr(self, "_cdp_mode", False):
            try:
                self.context.close()
            except Exception:
                pass
        if self.playwright:
            try:
                self.playwright.stop()
            except Exception:
                pass
        log("👋 浏览器已关闭")

    # ---------- 断点续传 ----------

    def get_completed_tags(self) -> set:
        """扫描已完成的标签（文件名含标签，且文件>5KB）。按当前产品名前缀匹配。"""
        done = set()
        prefix = f"{self.product.name}_达人_"
        for f in Config.OUTPUT_DIR.glob(f"{prefix}*_*.xlsx"):
            if f.stat().st_size < 5000:
                continue
            stem = f.stem.replace(prefix, "")
            # 文件名格式：标签_YYYYMMDD_HHMMSS
            # 时间戳 = 8位日期 + 1下划线 + 6位时间 = 15字符
            if len(stem) > 16 and stem[-16] == '_':
                tag = stem[:-16]
                done.add(tag)
        return done

    # ---------- 数据解析 ----------

    def _parse_item(self, item: dict, tag: str) -> dict:
        """将API返回的单条达人数据解析为标准化字典"""
        vi = item.get("viewInfo") or {}
        nickname = str(item.get("nickname", ""))
        pid = str(item.get("promoterId", ""))

        # 粉丝数
        fans = vi.get("PromoterFansCount") or item.get("fansNum", "")
        if isinstance(fans, (int, float)):
            fans_raw = fans
            fans_display = f"{fans/10000:.1f}万" if fans >= 10000 else str(int(fans))
        else:
            try:
                fans_raw = float(str(fans).replace("万", "")) * 10000
            except Exception:
                fans_raw = 0
            fans_display = str(fans)

        # 带货品类
        hs = item.get("hotSaleChannel", "")
        if isinstance(hs, list):
            hs = "/".join(hs)

        # 热卖商品
        ti = item.get("topItemList", [])
        if isinstance(ti, list):
            tit = " | ".join([
                t.get("itemName", "") for t in ti if t.get("itemName")
            ][:5])
        else:
            tit = ""

        # 佣金率
        comm = item.get("minCommissionRate", "")
        if isinstance(comm, (int, float)):
            comm_display = f"{comm/100:.1f}%"
        else:
            comm_display = str(comm)

        # 客单价
        ap = item.get("promoteAvgPrice", "")
        if isinstance(ap, (int, float)):
            ap_display = f"{ap/100:.2f}"
        else:
            ap_display = str(ap)

        # 性别
        g = item.get("gender", "")
        gd = {"M": "男", "F": "女"}.get(str(g), str(g))

        return {
            "采集时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "内容标签": tag,
            "达人名称": nickname[:50],
            "昵称": nickname,
            "达人性别": gd,
            "快手ID": pid,
            "粉丝数": fans_display,
            "粉丝数_raw": fans_raw,
            "带货品类": str(hs)[:80],
            "top_items_text": tit,
            "top_items_brief": tit[:60] + ("..." if len(tit) > 60 else ""),
            "场均销售额": str(vi.get("PromoterLiveAvgGMV30d", "")),
            "场均观看": str(vi.get("PromoterLiveAvgVisitorNum30d", "")),
            "带货评分": str(vi.get("PromoterScore", "")),
            "佣金率": comm_display,
            "客单价": ap_display,
            "有联系方式": "是",
            "达人地域": "",       # v11.2: 达人所在省市
            "粉丝年龄": "",       # v11.2: 粉丝年龄分布
            "粉丝性别": "",       # v11.2: 粉丝性别分布
            "粉丝城市": "",       # v11.2: 粉丝城市等级分布
            "产品契合度": 0,
            "契合等级": "",
            self.product.assoc_field: "",
            self.product.audience_label: "",
            "评分明细": "",
            "备注": f"佣金率:{comm_display} | 客单价:{ap_display}",
        }

    # ---------- API采集 ----------

    def fetch_tag(self, tag: str, max_retry: int = 8) -> list | None:
        """
        按内容标签拉取该标签下全部达人（有联系方式）。
        返回达人数据列表；登录过期返回 None。
        """
        if tag not in Config.TAG_IDS:
            log(f"  ⛔ 未知标签ID: {tag}，跳过")
            return []

        tag_ids = Config.TAG_IDS[tag]

        # v11.5+: API级筛选参数生效时，直接复用构造好的请求参数
        # （基于CDP抓包验证的真实映射，平台返回筛选后数据）
        page_params = getattr(self, "_page_filter_params", None)
        if page_params:
            params_base = dict(page_params)
            # 每个标签强制替换contentTag（筛选参数保留）
            params_base["contentTag"] = tag_ids
            params_base["offset"] = 0
            params_base["limit"] = self.page_size
            log(f"  📥 [{tag}] 开始拉取数据（API级筛选参数）...")
            log(f"  🔍 平台筛选已生效，仅拉取符合条件的达人")
        else:
            params_base = {
                "orderField": 0,
                "orderType": 1,
                "limit": self.page_size,
                "offset": 0,
                "type": self.promoter_type,
                "contentTag": tag_ids,
                "hotSaleChannelIdList": [],
                "hotSaleSubChannelId": [],
            }
            if self.has_contact:
                params_base["hasContact"] = 1

            # 尝试在API级别添加性别筛选（减少返回数据量）
            fc = self.filter_config
            self._api_gender_tried = False
            self._api_gender_failed = False
            if fc.get("gender"):
                gender_map = {"男": "M", "女": "F"}
                api_genders = [gender_map.get(g, g) for g in fc["gender"] if gender_map.get(g)]
                if len(api_genders) == 1:
                    params_base["gender"] = api_genders[0]
                    self._api_gender_tried = True

            log(f"  📥 [{tag}] 开始拉取数据...")
            type_names = {0: "全部达人", 1: "直播达人", 2: "视频达人"}
            filter_parts = [f"类型: {type_names.get(self.promoter_type, '未知')}"]
            filter_parts.append(f"联系方式: {'仅看有联系方式' if self.has_contact else '不过滤'}")
            if self._api_gender_tried:
                filter_parts.append(f"性别: {params_base['gender']}（API级）")
            elif fc.get("gender"):
                filter_parts.append(f"性别: {','.join(fc['gender'])}（本地筛选）")
            log(f"  🔍 {' | '.join(filter_parts)}")

        # 分页状态初始化（两种模式共用）
        all_items = []
        offset = 0
        total = None
        page_num = 0
        retry = 0
        self._tag_start_time = time.time()

        # 每3秒发一个心跳点，证明脚本活着
        last_heartbeat = time.time()

        while True:
            # 停止检查
            if self._stop:
                log(f"\n  🛑 [{tag}] 用户停止，已采 {len(all_items)} 条")
                break

            # 心跳：超过3秒没输出就打个点
            if time.time() - last_heartbeat > 3:
                print(".", end="", flush=True)
                last_heartbeat = time.time()

            params = params_base.copy()
            params["offset"] = offset

            try:
                result = self.page.evaluate("""
                    async ({url, body}) => {
                        const r = await fetch(url, {
                            method: 'POST',
                            headers: {
                                'Content-Type': 'application/json',
                                'Accept': 'application/json'
                            },
                            body: JSON.stringify(body),
                            credentials: 'include'
                        });
                        return await r.text();
                    }
                """, {"url": Config.API_URL, "body": params})

                data = json.loads(result)

                if data.get("result") != 1:
                    err_msg = str(data.get("error_msg", ""))
                    # 如果API不支持gender参数，降级为本地筛选
                    if self._api_gender_tried and not self._api_gender_failed:
                        self._api_gender_failed = True
                        params_base.pop("gender", None)
                        log(f"\n  ⚠️ API不支持性别筛选参数，降级为本地预筛选")
                        continue
                    if "操作过快" in err_msg or "频繁" in err_msg:
                        retry += 1
                        if retry > max_retry:
                            log(f"\n  ⚠️ 连续限流 {max_retry} 次，已采 {len(all_items)} 条，放弃")
                            break
                        wait = min(60, 3 * (2 ** retry))
                        log(f"\n  ⏳ 限流保护，等待 {wait} 秒 (第{retry}/{max_retry}次)...")
                        time.sleep(wait)
                        last_heartbeat = time.time()
                        continue
                    if "登录" in err_msg or "失效" in err_msg:
                        log(f"\n  🔐 登录过期: {err_msg}")
                        return None
                    log(f"\n  ❌ API错误: {err_msg}")
                    break

                retry = 0
                plist = data.get("data", {}).get("promoterList", [])

                if total is None:
                    total = data.get("data", {}).get("total", 0)
                    # 如果API级性别筛选返回0结果，降级重试
                    if total == 0 and self._api_gender_tried and not self._api_gender_failed:
                        self._api_gender_failed = True
                        params_base.pop("gender", None)
                        offset = 0
                        log(f"  ⚠️ API级性别筛选返回0结果，降级为本地预筛选")
                        continue
                    log(f"  📊 [{tag}] 共 {total} 人，预计 {max(1, total // self.page_size)} 页")

                if not plist:
                    break

                for item in plist:
                    all_items.append(self._parse_item(item, tag))

                page_num += 1
                offset += self.page_size
                last_heartbeat = time.time()

                # 每页都报进度
                pct = min(100, round(offset / total * 100)) if total else 0
                elapsed = int(time.time() - self._tag_start_time)
                log(f"  📄 [{tag}] 第{page_num}页 | {len(all_items)}/{total}人 ({pct}%) | 耗时{elapsed}秒")

                # GUI 进度回调
                if self._on_progress:
                    try:
                        self._on_progress(tag, len(all_items), total or 0, page_num)
                    except Exception:
                        pass

                if offset >= total:
                    break

                time.sleep(self.page_delay)

            except Exception as e:
                retry += 1
                if retry > max_retry:
                    log(f"\n  ❌ 重试耗尽: {e}")
                    break
                wait = min(30, retry * 5)
                log(f"\n  ⚠️ 异常: {e}，{wait}秒后重试 (第{retry}/{max_retry}次)")
                time.sleep(wait)
                last_heartbeat = time.time()
                continue

        elapsed_tag = int(time.time() - self._tag_start_time)
        log(f"  ✅ [{tag}] 完成！{len(all_items)} 条 | 总耗时 {elapsed_tag} 秒")
        return all_items

    # ---------- 详情 + 粉丝画像补充 ----------

    def enrich_detail_fans(self, dlist: list, tag: str) -> None:
        """批量查询详情API（达人地域）和粉丝画像API（年龄/性别/城市）"""
        if not dlist:
            return
        total = len(dlist)
        log(f"  📋 [{tag}] 查询 {total} 人详情+粉丝画像...")
        checked_detail = 0
        checked_fans = 0

        for i, d in enumerate(dlist):
            pid = d["快手ID"]
            # 查询详情API → 达人地域
            try:
                result = self.page.evaluate("""
                    async ({url, promoterId}) => {
                        const res = await fetch(url + '?promoterId=' + promoterId + '&type=1', {
                            headers: {'Accept': 'application/json'}, credentials: 'include'
                        });
                        return await res.text();
                    }
                """, {"url": Config.API_DETAIL_URL, "promoterId": pid})
                data = json.loads(result)
                detail = data.get("data", {}) if data.get("result") == 1 else {}
                addr = detail.get("addressInfo", {})
                province = addr.get("province", "")
                city = addr.get("city", "")
                if province or city:
                    d["达人地域"] = f"{province}{city}" if province else city
                checked_detail += 1
            except Exception:
                pass

            # 查询粉丝画像API
            try:
                result = self.page.evaluate("""
                    async ({url, promoterId}) => {
                        const res = await fetch(url + '?promoterId=' + promoterId + '&timeRangeType=3', {
                            headers: {'Accept': 'application/json'}, credentials: 'include'
                        });
                        return await res.text();
                    }
                """, {"url": Config.API_FANS_URL, "promoterId": pid})
                data = json.loads(result)
                fans = data.get("data", {}) if data.get("result") == 1 else {}

                # 粉丝年龄
                age_list = fans.get("fansAgeFeature", [])
                if age_list:
                    parts = []
                    for a in sorted(age_list, key=lambda x: x.get("ageFeatureRate", 0), reverse=True)[:4]:
                        parts.append(f"{a.get('ageFeature','')} {a.get('ageFeatureRate',0)/100:.1f}%")
                    d["粉丝年龄"] = " | ".join(parts)

                # 粉丝性别
                gender_list = fans.get("fansGenderFeature", [])
                if gender_list:
                    parts = []
                    for g in sorted(gender_list, key=lambda x: x.get("genderFeatureRate", 0), reverse=True):
                        parts.append(f"{g.get('genderFeature','')} {g.get('genderFeatureRate',0)/100:.1f}%")
                    d["粉丝性别"] = " | ".join(parts)

                # 粉丝城市等级
                city_list = fans.get("fansCityLevelFeature", [])
                if city_list:
                    parts = []
                    for c in sorted(city_list, key=lambda x: x.get("cityLevelRate", 0), reverse=True)[:4]:
                        parts.append(f"{c.get('cityLevelFeature','')} {c.get('cityLevelRate',0)/100:.1f}%")
                    d["粉丝城市"] = " | ".join(parts)

                checked_fans += 1
            except Exception:
                pass

            if (i + 1) % 50 == 0:
                log(f"  [{tag}] 已查询 {i + 1}/{total}...")
            time.sleep(0.25)

        log(f"  📋 [{tag}] 详情:{checked_detail}/{total} | 粉丝画像:{checked_fans}/{total}")

    # ---------- API级筛选参数（v11.5+：基于CDP抓包验证的真实映射） ----------
    # 2026-08-14 通过 CDP 接管用户已登录 Chrome 抓取达人广场 promoter/list 请求体验证
    # 不再依赖页面 UI 点击，直接构造 API 请求体参数

    # 粉丝年龄：fansAgeFeature（数组，1-6）
    # 页面选项：18岁以下/18-23岁/24-30岁/31-40岁/41-50岁/50岁以上
    _API_AGE_MAP = {
        "18岁以下": 1, "18~23岁": 2, "18-23岁": 2,
        "24~30岁": 3, "24-30岁": 3,
        "31~40岁": 4, "31-40岁": 4,
        "41~50岁": 5, "41-50岁": 5,
        "50岁以上": 6, "50岁+": 6,
    }

    # 粉丝性别：fansGenderFeature（单个数字，1=男粉丝多 2=女粉丝多）
    # 页面选项：男粉丝多/女粉丝多（注意不是"女多男少"）
    _API_FAN_GENDER_MAP = {
        "男性为主": 1, "男粉丝多": 1,
        "女性为主": 2, "女粉丝多": 2,
    }

    # 达人性别：gender（字符串 "M"/"F"）
    _API_SELLER_GENDER_MAP = {"男": "M", "男性": "M", "女": "F", "女性": "F"}

    # 粉丝城市划分：fansAddressFeature（数组，1-5）
    # 页面选项：一线城市/二线城市/三线城市/四线城市/五线城市及以下
    # 注意：页面无"新一线"选项，新一线城市归入一级(1)
    _API_CITY_MAP = {
        "一线城市": 1, "新一线城市": 1, "新一线": 1,
        "二线城市": 2, "二线": 2,
        "三线城市": 3, "三线": 3,
        "四线城市": 4, "四线": 4,
        "五线城市": 5, "五线": 5, "五线城市及以下": 5, "四线及以下": 4,
    }

    # 达人地域：promoterProvinceCode（标准国标省份码）+ promoterCountryCode="86"
    _API_PROVINCE_CODE = {
        "北京": "110000", "天津": "120000", "河北": "130000", "山西": "140000",
        "内蒙古": "150000", "辽宁": "210000", "吉林": "220000", "黑龙江": "230000",
        "上海": "310000", "江苏": "320000", "浙江": "330000", "安徽": "340000",
        "福建": "350000", "江西": "360000", "山东": "370000", "河南": "410000",
        "湖北": "420000", "湖南": "430000", "广东": "440000", "广西": "450000",
        "海南": "460000", "重庆": "500000", "四川": "510000", "贵州": "520000",
        "云南": "530000", "西藏": "540000", "陕西": "610000", "甘肃": "620000",
        "青海": "630000", "宁夏": "640000", "新疆": "650000",
        "台湾": "710000", "香港": "810000", "澳门": "820000",
    }

    # ---------- 页面级筛选（v11.5: 先筛后拉，兼容降级） ----------
    # GUI筛选值 → 达人广场页面选项值 映射（降级到页面点击时使用）
    _PAGE_FILTER_VALUE_MAP = {
        "fan_age": {
            "18~23岁": "18-23岁", "24~30岁": "24-30岁",
            "31~40岁": "31-40岁", "41~50岁": "41-50岁",
            "50岁以上": "50岁以上", "18岁以下": "18岁以下",
        },
        "fan_gender": {
            "女性为主": "女粉丝多", "男性为主": "男粉丝多",
        },
        "fan_city": {
            "四线城市": "四线城市", "五线城市": "五线城市及以下",
            "新一线城市": "一线城市",
        },
    }

    # 筛选维度别名（页面上的文字）
    _PAGE_FILTER_DIM_ALIASES = {
        "gender": ["达人性别", "性别"],
        "region": ["达人地域", "地域", "所在地域", "地区"],
        "fan_age": ["粉丝年龄", "年龄", "粉丝年龄分布"],
        "fan_gender": ["粉丝性别", "粉丝性别分布", "性别分布"],
        "fan_city": ["粉丝城市划分", "城市划分", "粉丝城市", "城市等级"],
    }

    def _find_clickable(self, text: str, exact: bool = True):
        """查找页面上可见的文本元素（可点击）"""
        for e in (exact, not exact):
            try:
                loc = self.page.get_by_text(text, exact=e).first
                if loc.count() > 0 and loc.is_visible():
                    return loc
            except Exception:
                pass
        return None

    def _click_filter_dimension(self, dim_key: str) -> bool:
        """点击筛选维度名（展开下拉/选项区）"""
        for alias in self._PAGE_FILTER_DIM_ALIASES.get(dim_key, [dim_key]):
            loc = self._find_clickable(alias)
            if loc:
                try:
                    loc.click(timeout=3000)
                    time.sleep(0.8)
                    return True
                except Exception:
                    pass
        return False

    def _click_filter_option(self, dim_key: str, value: str) -> bool:
        """点击筛选选项值（含GUI→页面映射，多策略匹配）"""
        mapped = self._PAGE_FILTER_VALUE_MAP.get(dim_key, {}).get(value, value)
        candidates = [mapped, value]
        short = mapped.replace("岁", "").replace("城市", "")
        if short != mapped:
            candidates.append(short)
        for cand in candidates:
            for e in (True, False):
                loc = self._find_clickable(cand, exact=e)
                if loc:
                    try:
                        loc.click(timeout=3000)
                        time.sleep(0.5)
                        return True
                    except Exception:
                        pass
        return False

    def _capture_list_params(self, wait_seconds: float = 4.0) -> dict | None:
        """监听页面自动发出的promoter/list请求，捕获请求体参数

        筛选点击后页面会自动重新请求列表，此时捕获的请求体
        就是快手自己构造的"完整筛选参数"，可直接复用翻页
        """
        captured = []

        def _on_request(request):
            try:
                if "promoter/list" in request.url and request.method == "POST":
                    body = json.loads(request.post_data or "{}")
                    captured.append(body)
            except Exception:
                pass

        self.page.on("request", _on_request)
        try:
            time.sleep(wait_seconds)
        finally:
            self.page.remove_listener("request", _on_request)

        if captured:
            return captured[-1]  # 最后一次请求 = 筛选全部设置完后的状态
        return None

    def _click_tag_on_page(self, tag: str) -> bool:
        """点击页面"内容标签"区的当前标签（页面单选，点击即切换）"""
        # 先尝试展开内容标签区（如果页面有折叠）
        for alias in ["内容标签", "全选"]:
            loc = self._find_clickable(alias)
            if loc:
                try:
                    loc.click(timeout=3000)
                    time.sleep(0.5)
                except Exception:
                    pass
                break
        loc = self._find_clickable(tag)
        if loc:
            try:
                loc.click(timeout=3000)
                time.sleep(1.5)
                return True
            except Exception:
                pass
        return False

    def apply_page_filters(self, tag: str) -> dict | None:
        """v11.5+：直接构造API请求参数（平台先筛后拉，无需页面点击）

        基于 2026-08-14 CDP 抓包验证的真实参数映射：
          - fansAgeFeature: [1-6]      粉丝年龄
          - fansGenderFeature: 1/2     粉丝性别（1=男粉丝多 2=女粉丝多）
          - gender: "M"/"F"            达人性别
          - fansAddressFeature: [1-5]  粉丝城市划分
          - promoterProvinceCode: 国标省份码 + promoterCountryCode="86"
          - hasContact: 1              有联系方式

        返回构造好的请求参数dict（供fetch_tag翻页复用）；
        无筛选条件时返回None
        """
        fc = self.filter_config
        active = {k: v for k, v in fc.items() if v}
        if not active and not self.has_contact:
            return None  # 无筛选条件，不需要API级筛选

        tag_ids = Config.TAG_IDS.get(tag, [])
        params = {
            "orderField": 0,
            "orderType": 1,
            "limit": self.page_size,
            "offset": 0,
            "type": self.promoter_type,
            "contentTag": tag_ids,
            "hotSaleChannelIdList": [],
            "hotSaleSubChannelId": [],
        }

        applied = []   # 已应用的API级筛选描述
        skipped = []   # 无法映射的项（本地筛选兜底）

        # 1. 有联系方式
        if self.has_contact:
            params["hasContact"] = 1
            applied.append("有联系方式")

        # 2. 达人性别（API仅支持单值，多选走本地筛选）
        if fc.get("gender"):
            api_g = [self._API_SELLER_GENDER_MAP.get(g) for g in fc["gender"]]
            api_g = [g for g in api_g if g]
            if len(api_g) == 1:
                params["gender"] = api_g[0]
                applied.append(f"达人性别={fc['gender'][0]}")
            elif len(api_g) > 1:
                skipped.append(f"达人性别多选({','.join(fc['gender'])})")

        # 3. 达人地域（省份码）
        if fc.get("region"):
            codes = [self._API_PROVINCE_CODE.get(r) for r in fc["region"]]
            valid_codes = [c for c in codes if c]
            if len(valid_codes) == 1:
                params["promoterProvinceCode"] = valid_codes[0]
                params["promoterCountryCode"] = "86"
                applied.append(f"地域={fc['region'][0]}")
            elif len(valid_codes) > 1:
                # API疑似仅支持单省份：逐省份轮采由上层处理，此处取第一个并提示
                params["promoterProvinceCode"] = valid_codes[0]
                params["promoterCountryCode"] = "86"
                applied.append(f"地域={fc['region'][0]}(首省)")
                skipped.append(f"地域多选仅生效首省: {','.join(fc['region'])}")
            else:
                skipped.append(f"地域码缺失: {','.join(fc['region'])}")

        # 4. 粉丝年龄（数组）
        if fc.get("fan_age"):
            ages = sorted({self._API_AGE_MAP[a] for a in fc["fan_age"] if a in self._API_AGE_MAP})
            if ages:
                params["fansAgeFeature"] = ages
                applied.append(f"粉丝年龄={','.join(fc['fan_age'])}")
            unmapped = [a for a in fc["fan_age"] if a not in self._API_AGE_MAP]
            if unmapped:
                skipped.append(f"年龄未映射: {','.join(unmapped)}")

        # 5. 粉丝性别（单值：1=男粉丝多 2=女粉丝多）
        if fc.get("fan_gender"):
            vals = {self._API_FAN_GENDER_MAP.get(g) for g in fc["fan_gender"]}
            vals.discard(None)
            if len(vals) == 1:
                params["fansGenderFeature"] = vals.pop()
                applied.append(f"粉丝性别={','.join(fc['fan_gender'])}")
            else:
                skipped.append(f"粉丝性别多选/未映射: {','.join(fc['fan_gender'])}")

        # 6. 粉丝城市划分（数组）
        if fc.get("fan_city"):
            cities = sorted({self._API_CITY_MAP[c] for c in fc["fan_city"] if c in self._API_CITY_MAP})
            if cities:
                params["fansAddressFeature"] = cities
                applied.append(f"城市划分={','.join(fc['fan_city'])}")
            unmapped = [c for c in fc["fan_city"] if c not in self._API_CITY_MAP]
            if unmapped:
                skipped.append(f"城市未映射: {','.join(unmapped)}")

        # 保存供 fetch_tag 翻页复用
        self._page_filter_params = params
        self._page_filter_active = True
        self._api_gender_tried = "gender" in params
        self._api_gender_failed = False

        log(f"  🎛️ [{tag}] API级筛选参数已构造（{len(applied)}项）")
        for a in applied:
            log(f"    ✅ {a}")
        for s in skipped:
            log(f"    ⚠️ {s} → 本地筛选兜底")
        param_show = {k: v for k, v in params.items()
                      if k not in ("hotSaleChannelIdList", "hotSaleSubChannelId")}
        log(f"  🔍 参数: {json.dumps(param_show, ensure_ascii=False)[:300]}")

        return params

    # ---------- 预筛选（查询详情前） ----------

    def pre_filter(self, dlist: list, tag: str) -> list:
        """预筛选：在查询详情API前，先用列表API已有的数据过滤
        
        列表API响应已包含：性别(gender)、带货品类(hotSaleChannel)、热卖商品(topItemList)
        可在此阶段过滤：
          - 性别（列表API已包含gender字段）
          - 产品关联度（从top_items_text匹配关键词）
        不能在此阶段过滤（需查详情/粉丝画像API）：
          - 达人地域（需详情API）
          - 粉丝年龄/性别/城市（需粉丝画像API）
        
        好处：大幅减少enrich_detail_fans的API调用量
        例如：697人 → 预筛选后60人 → 只查60人详情（原来查697人）
        """
        if not dlist:
            return dlist

        original_count = len(dlist)
        fc = self.filter_config

        # 1. 先评分（关联度）—— 评分用的字段都在列表API响应中
        af = self.product.assoc_field
        al = self.product.audience_label
        tg = self.product.target_gender
        assoc_on = self.product.has_assoc_rules
        for d in dlist:
            score, detail, matched_kws = Scorer.score(d, tag, self.product)
            d["产品契合度"] = score
            d["契合等级"] = Scorer.level(score, self.product)
            if assoc_on:
                d[af] = " / ".join(matched_kws) if matched_kws else "无"
            else:
                d[af] = "—"  # 通用手动模式：关联匹配不参与
            gd = d.get("达人性别")
            if tg == "无":
                aud = "中"
            elif gd == tg:
                aud = "高" if score >= 60 else "中"
            else:
                aud = "低"
            d[al] = aud
            d["评分明细"] = detail

        # 2. 按性别预筛选（列表API已包含gender字段，无需查详情）
        if fc.get("gender") and not self._api_gender_failed:
            # API级性别筛选已生效，不需要本地再筛
            pass
        elif fc.get("gender"):
            before = len(dlist)
            dlist = [d for d in dlist if not d.get("达人性别") or d.get("达人性别") in fc["gender"]]
            if before != len(dlist):
                log(f"  ⚡ [{tag}] 性别预筛选: {before} -> {len(dlist)} 人"
                    f"（跳过{before - len(dlist)}人详情查询）")

        # 3. 按产品关联预筛选（无关联的直接排除，不查询详情）
        #    通用手动模式（未配置关联规则）下不做此项剔除
        if self.product.has_assoc_rules:
            before_wl = len(dlist)
            dlist = [d for d in dlist if d.get(af, "无") != "无"]
            if before_wl != len(dlist):
                log(f"  ⚡ [{tag}] {af}预筛选: {before_wl} -> {len(dlist)} 人"
                    f"（排除{before_wl - len(dlist)}人无关联，跳过详情查询）")
        else:
            log(f"  ⚡ [{tag}] 通用手动模式：跳过关联剔除（仅按销售额门槛+手动标签筛选）")

        if not dlist:
            log(f"  ⚡ [{tag}] 预筛选后无达人（{original_count} -> 0）")
        else:
            reduction = (1 - len(dlist) / original_count) * 100
            log(f"  ⚡ [{tag}] 预筛选完成: {original_count} -> {len(dlist)} 人"
                f"（减少{reduction:.0f}%详情查询）")

        return dlist

    # ---------- 采集后筛选（需详情数据） ----------

    def apply_filters(self, dlist: list, tag: str) -> list:
        """根据 GUI 筛选条件过滤达人（采集后过滤）"""
        fc = self.filter_config
        if not fc or all(not v for v in fc.values()):
            return dlist

        before = len(dlist)
        result = []
        reject_samples = []  # 记录前3个被筛原因供调试

        for d in dlist:
            keep = True
            reason = ""

            # 达人性别（精确匹配）
            if keep and fc.get("gender"):
                gender = d.get("达人性别", "")
                if gender and gender not in fc["gender"]:
                    keep = False
                    reason = f"性别={gender}"

            # 达人地域（省市模糊匹配，API 级 province 预筛 + 本地兜底）
            if keep and fc.get("region"):
                region = d.get("达人地域", "")
                if region and not any(r in region for r in fc["region"]):
                    keep = False
                    reason = f"地域={region}"

            # 地级市（本地匹配达人地域列，需采集粉丝画像；无画像列则跳过，不误删）
            if keep and fc.get("city"):
                region = d.get("达人地域", "")
                if region and not any(c in region for c in fc["city"]):
                    keep = False
                    reason = f"城市={region}"

            # 粉丝年龄（API返回"24~30"无"岁"，选项"24~30岁"带"岁"，去"岁"后比较）
            if keep and fc.get("fan_age"):
                age_str = d.get("粉丝年龄", "")
                if age_str:
                    main = age_str.split(" | ")[0].rsplit(" ", 1)[0]
                    # 去掉"岁"后缀再比较，兼容"24~30" vs "24~30岁"
                    main_clean = main.rstrip("岁")
                    fan_ages_clean = [a.rstrip("岁") for a in fc["fan_age"]]
                    if main_clean not in fan_ages_clean:
                        keep = False
                        reason = f"年龄={main}"

            # 粉丝性别（主导性别含"女"/"男"即可）
            if keep and fc.get("fan_gender"):
                fg_str = d.get("粉丝性别", "")
                if fg_str:
                    first_label = fg_str.split(" | ")[0].split(" ")[0]
                    opts = fc["fan_gender"]
                    ok = ("女性为主" in opts and "女" in first_label) or \
                         ("男性为主" in opts and "男" in first_label)
                    if not ok:
                        keep = False
                        reason = f"粉性={first_label}"

            # 粉丝城市等级（API可能返回"一线"/"一线城市"，用包含匹配）
            if keep and fc.get("fan_city"):
                city_str = d.get("粉丝城市", "")
                if city_str:
                    main = city_str.split(" | ")[0].rsplit(" ", 1)[0]
                    # 双向包含匹配：API"一线"⊂选项"一线城市" 或 选项"一线"⊂API"一线城市"
                    matched = False
                    for c in fc["fan_city"]:
                        if main in c or c in main:
                            matched = True
                            break
                    if not matched:
                        keep = False
                        reason = f"城级={main}"

            if keep:
                result.append(d)
            elif len(reject_samples) < 3:
                reject_samples.append(f"{d.get('昵称','?')[:12]} → {reason}")

        after = len(result)
        if before != after:
            log(f"  🔍 [{tag}] 筛选: {before} → {after} 人")
            if reject_samples:
                log(f"     💡 示例被筛: {' | '.join(reject_samples)}")
        return result

    # ---------- 分销销售额(近30日)过滤 ----------

    def filter_low_dist_sales(self, dlist: list, tag: str) -> list:
        """近30日分销销售额(场均GMV)低于阈值的达人不采集。

        字段来源：promoter/list 列表项 viewInfo.PromoterLiveAvgGMV30d，
        解析后存于 d["场均销售额"]（页面文案"直播场均销售额/30日直播均场销售额"）。
        在 enrich_detail_fans 之前调用，避免对不达标达人浪费详情/粉丝画像 API 查询。
        返回保留的达人列表。
        """
        thr = self.product.min_dist_sales_30d
        if not thr or thr <= 0:
            return dlist

        before = len(dlist)
        kept = []
        dropped = 0
        samples = []

        for d in dlist:
            raw = d.get("场均销售额", "")
            try:
                val = float(raw) if raw not in ("", None) else None
            except (ValueError, TypeError):
                val = None

            if val is None:
                # 无数据：按 DROP_NO_SALES_DATA 决定保留或剔除
                if self.product.drop_no_sales_data:
                    dropped += 1
                    if len(samples) < 3:
                        samples.append(f"{d.get('昵称','?')[:12]} 场均销售额=无数据")
                    continue
                kept.append(d)
                continue

            if val < thr:
                dropped += 1
                if len(samples) < 3:
                    samples.append(f"{d.get('昵称','?')[:12]} 场均销售额={val}")
            else:
                kept.append(d)

        if dropped:
            log(f"  💰 [{tag}] 近30日分销销售额<{thr} 过滤: {before} → {len(kept)} 人 (剔除{dropped})")
            if samples:
                log(f"     💡 示例剔除: {' | '.join(samples)}")
        elif before != len(kept):
            log(f"  💰 [{tag}] 近30日分销销售额过滤: {before} → {len(kept)} 人")
        return kept

    # ---------- 评分 + 保存 ----------

    def score_and_save(self, dlist: list, tag: str) -> Path | None:
        """评分 → 排序 → 保存Excel，返回文件路径"""
        if not dlist:
            log(f"  [{tag}] 无数据，跳过保存")
            return None

        # 检查是否已在pre_filter阶段评分过
        already_scored = any(d.get("产品契合度", 0) > 0 for d in dlist)

        if not already_scored:
            # 旧路径：未经过pre_filter，在这里评分+过滤
            log(f"  📊 [{tag}] 正在评分（{self.product.name}人群匹配）...")
            af = self.product.assoc_field
            al = self.product.audience_label
            tg = self.product.target_gender
            assoc_on = self.product.has_assoc_rules
            for d in dlist:
                score, detail, matched_kws = Scorer.score(d, tag, self.product)
                d["产品契合度"] = score
                d["契合等级"] = Scorer.level(score, self.product)
                if assoc_on:
                    d[af] = " / ".join(matched_kws) if matched_kws else "无"
                else:
                    d[af] = "—"  # 通用手动模式：关联匹配不参与
                gd = d.get("达人性别")
                if tg == "无":
                    aud = "中"
                elif gd == tg:
                    aud = "高" if score >= 60 else "中"
                else:
                    aud = "低"
                d[al] = aud
                d["评分明细"] = detail

            # 只保留有产品关联的达人，无关联的直接排除
            # 通用手动模式（未配置关联规则）下保留全部，由用户手动选标签决定
            if self.product.has_assoc_rules:
                original = len(dlist)
                dlist[:] = [d for d in dlist if d.get(af, "无") != "无"]
                filtered = original - len(dlist)
                if filtered > 0:
                    log(f"  🔍 [{tag}] 排除无{af}达人:{filtered}人 | 保留:{len(dlist)}人")
                if not dlist:
                    log(f"  [{tag}] 过滤后无达人，跳过保存")
                    return None
        else:
            log(f"  📊 [{tag}] 评分完成（预筛选阶段已处理）")

        # 按契合度降序 → 粉丝数降序
        dlist.sort(key=lambda x: (-x["产品契合度"], -x.get("粉丝数_raw", 0)))

        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_tag = tag.replace("/", "_").replace("\\", "_")
        filepath = Config.OUTPUT_DIR / f"{self.product.name}_达人_{safe_tag}_{timestamp}.xlsx"

        log(f"  💾 [{tag}] 正在保存Excel...")
        wb = Workbook()
        ws = wb.active
        ws.title = f"{tag}达人"

        headers = [
            "序号", "产品契合度", "契合等级", self.product.assoc_field, "内容标签", "达人性别",
            "达人地域", "粉丝年龄", "粉丝性别", "粉丝城市", self.product.audience_label,
            "联系方式", "达人名称", "快手ID", "粉丝数", "带货品类",
            "场均销售额", "场均观看", "带货评分", "佣金率", "客单价",
            "商品示例", "评分明细", "采集时间", "备注",
        ]

        # 样式
        hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hfn = Font(color="FFFFFF", bold=True, size=11)
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        light_green = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        yellow = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")
        red = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hf
            cell.font = hfn
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for ri, d in enumerate(dlist, 2):
            vals = [
                ri - 1, d["产品契合度"], d["契合等级"], d[self.product.assoc_field], d["内容标签"],
                d["达人性别"], d["达人地域"], d["粉丝年龄"], d["粉丝性别"], d["粉丝城市"],
                d[self.product.audience_label], d["有联系方式"], d["达人名称"],
                d["快手ID"], d["粉丝数"], d["带货品类"], d["场均销售额"],
                d["场均观看"], d["带货评分"], d["佣金率"], d["客单价"],
                d["top_items_brief"], d["评分明细"], d["采集时间"], d["备注"],
            ]
            score = d["产品契合度"]
            has_wl = d.get(self.product.assoc_field, "无") != "无"
            fill = (
                green if score >= 75
                else light_green if score >= 60
                else yellow if score >= 45
                else PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") if has_wl
                else red
            )
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=c, value=v)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        col_widths = [5, 8, 14, 24, 8, 8, 10, 22, 14, 22, 8, 8, 18, 18, 10, 28, 14, 12, 10, 8, 8, 22, 30, 16, 30]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        ws.freeze_panes = "A2"
        Config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)

        hi = sum(1 for d in dlist if d["产品契合度"] >= 75)
        mi = sum(1 for d in dlist if 60 <= d["产品契合度"] < 75)
        fe = sum(1 for d in dlist if d.get("达人性别") == "女")
        log(f"  💾 [{tag}] 已保存: {filepath.name}")
        log(f"  📈 [{tag}] 统计: ≥75分 {hi}人 | 60-74分 {mi}人 | 共{len(dlist)}人(全部{self.product.assoc_field}) | {self.product.audience_label} {fe}人")

        return filepath

    # ---------- 主流程 ----------

    def run(self, tags: list) -> dict:
        """
        主执行流程：登录 → 逐标签采集 → 评分保存 → 汇总
        返回 {tag: count} 的结果字典
        """
        completed = self.get_completed_tags()
        remaining = [t for t in tags if t not in completed]

        log("=" * 60)
        log(f"  🎯 {self.product.name} 快手达人采集工具 v11.7（多产品配置化 + 页面级先筛后拉 + 分销销售额过滤）")
        log(f"  📋 全部标签: {len(tags)} 个")
        log(f"  ✅ 已完成: {len(completed)} 个")
        log(f"  📌 待采集: {len(remaining)} 个")
        log("=" * 60)

        if completed:
            log(f"  已完成: {', '.join(sorted(completed))}")
        if not remaining:
            log("🎉 所有标签已采集完毕！")
            return {}

        log(f"  待采集: {', '.join(remaining)}")
        log("")

        # 启动浏览器
        self.start_browser()
        self.navigate_to_daren_square()

        # 登录
        if not self.wait_for_login():
            log("❌ 登录失败，退出")
            self.close()
            return {}

        # 逐标签采集
        results = {}
        all_files = []
        run_start = time.time()

        for idx, tag in enumerate(remaining, 1):
            # 停止检查
            if self._stop:
                log("🛑 用户请求停止，终止后续标签")
                break

            log("")
            log("━" * 50)
            log(f"  📌 [{idx}/{len(remaining)}] 正在处理标签: {tag}")
            log("━" * 50)

            # v11.5: 先做页面级筛选（平台先筛后拉）
            # 在页面上真实点击粉丝年龄/性别/地域/城市等筛选条件，
            # 拦截平台发出的请求参数，只拉取符合条件的达人
            page_params = self.apply_page_filters(tag)
            if page_params:
                self._page_filter_params = page_params
            else:
                self._page_filter_params = None

            dlist = self.fetch_tag(tag)

            if dlist is None:
                if self.relogin():
                    dlist = self.fetch_tag(tag)
                if dlist is None:
                    log(f"  ⛔ [{tag}] 登录失败，跳过")
                    continue

            if self._page_filter_params:
                # 页面级筛选已生效：平台已按条件过滤，无需本地预筛选
                # 但仍查详情/粉丝画像用于表格展示 + apply_filters 兜底
                if not dlist:
                    log(f"  [{tag}] 页面筛选后无达人，跳过")
                    continue
                log(f"  ⚡ [{tag}] 平台已筛选（跳过本地预筛选）")
            else:
                # 降级：本地预筛选（v11.4 逻辑，页面筛选未生效时兜底）
                # 性别+减肥关联在列表API响应中已有，无需查详情即可过滤
                dlist = self.pre_filter(dlist, tag)
                if not dlist:
                    log(f"  [{tag}] 预筛选后无达人，跳过")
                    continue

            # 近30日分销销售额过滤：低于阈值的不采集，放在详情查询之前以节省 API 调用
            dlist = self.filter_low_dist_sales(dlist, tag)
            if not dlist:
                log(f"  [{tag}] 分销销售额过滤后无达人，跳过")
                continue

            # v11.2: 查询详情 + 粉丝画像（只查询筛选后的达人）
            # need_portrait=False（通用极简模式）：跳过 per-达人画像查询，省 N×2 次接口；
            # 粉丝性别/年龄/地域的 API 级筛选已在 apply_page_filters 生效，不依赖此处
            if self.product.need_portrait:
                self.enrich_detail_fans(dlist, tag)
            else:
                log(f"  ⏩ [{tag}] 通用模式：跳过粉丝画像查询（need_portrait=False），"
                    f"仅按销售额门槛 + 手动标签筛选，省 {len(dlist)}×2 次接口")

            # 应用 GUI 筛选条件（采集后兜底过滤；页面已筛的基本全通过）
            dlist = self.apply_filters(dlist, tag)
            if not dlist:
                log(f"  [{tag}] 筛选后无达人，跳过保存")
                continue

            fp = self.score_and_save(dlist, tag)
            if fp:
                all_files.append(fp)
                results[tag] = len(dlist)

            # GUI 标签完成回调
            if self._on_tag_done:
                try:
                    self._on_tag_done(tag, len(dlist) if dlist else 0, str(fp) if fp else "")
                except Exception:
                    pass

            # 标签间延迟
            if idx < len(remaining) and not self._stop:
                log(f"  ⏸️ 标签间冷却 {self.tag_delay} 秒...")
                time.sleep(self.tag_delay)

        run_elapsed = int(time.time() - run_start)
        self.close()

        # 汇总
        self._print_summary(tags, completed, results, all_files, run_elapsed)

        # GUI 全部完成回调
        if self._on_all_done:
            try:
                self._on_all_done(results, all_files, run_elapsed)
            except Exception:
                pass

        return results

    def _print_summary(self, all_tags: list, completed: set, results: dict, files: list, elapsed: int):
        """打印采集汇总"""
        log("")
        log("=" * 60)
        log("  📊 采集任务完成！")
        log(f"  ⏱️ 总耗时: {elapsed} 秒")
        log("=" * 60)

        total_count = 0
        for tag in all_tags:
            if tag in completed:
                log(f"  ✅ {tag:<10s} (之前已完成)")
            elif tag in results:
                c = results[tag]
                total_count += c
                log(f"  ⭐ {tag:<10s}: {c:>4d} 人")
            else:
                log(f"  ❌ {tag:<10s}: 失败")

        log(f"\n  📦 本次新增: {len(results)} 个标签, {total_count} 条达人")
        log(f"  📄 生成文件: {len(files)} 个 Excel")
        for f in files:
            log(f"     → {f.name}")
        log("=" * 60)


# ============================================================
# 命令行接口
# ============================================================

def resolve_tags(args, product: "ProductConfig") -> list | None:
    """根据命令行参数确定要采集的标签列表"""
    if args.all:
        return Config.ALL_TAGS.copy()
    elif args.rec:
        return (product.recommended_tags or []).copy()
    elif args.tag:
        return [t.strip() for t in args.tag.split(",") if t.strip()]
    else:
        return interactive_select(product)


def interactive_select(product: "ProductConfig") -> list | None:
    """交互式选择标签"""
    log("")
    log("=" * 60)
    log(f"  {product.name} - 快手达人采集工具")
    log("=" * 60)

    rec = set(product.recommended_tags or [])
    print("\n📋 可选内容标签：")
    for i, tag in enumerate(Config.ALL_TAGS, 1):
        marker = "⭐" if tag in rec else "  "
        print(f"  {marker} {i:2d}. {tag}", end="")
        if i % 4 == 0:
            print()
    print()
    print(f"⭐ = 推荐标签（与「{product.name}」匹配度最高）")

    while True:
        print("\n请选择：")
        print("  1) 输入序号（逗号分隔），如: 10,15,7,6")
        print("  2) 输入 'rec' → 采集所有推荐标签")
        print("  3) 输入 'all' → 采集全部标签")
        print("  4) 输入标签名（逗号分隔），如: 健康,运动,美妆")
        print("  0) 退出")

        choice = input("\n> ").strip()
        if not choice:
            continue
        if choice == "0":
            return None
        if choice == "rec":
            return (product.recommended_tags or []).copy()
        if choice == "all":
            return Config.ALL_TAGS.copy()

        parts = [p.strip() for p in choice.split(",")]
        result = []
        valid = True
        for part in parts:
            if part.isdigit():
                idx = int(part)
                if 1 <= idx <= len(Config.ALL_TAGS):
                    result.append(Config.ALL_TAGS[idx - 1])
                else:
                    print(f"  无效序号: {idx}")
                    valid = False
                    break
            elif part in Config.ALL_TAGS:
                result.append(part)
            else:
                print(f"  未知标签: {part}")
                valid = False
                break

        if valid and result:
            return result
        print("  请重新输入")


def main():
    parser = argparse.ArgumentParser(
        description="快手达人采集工具（多产品配置化）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s --product 颜阿娇 --all         采集「颜阿娇」全部37个标签
  %(prog)s --product 颜阿娇 --rec         采集「颜阿娇」推荐标签
  %(prog)s --product 示例_母婴 --tag 亲子  采集「母婴」指定标签
  %(prog)s --product 示例_母婴 --rec       采集「母婴」推荐标签
  %(prog)s --tag 健康    采集默认产品的指定标签
  %(prog)s               交互模式（先选产品再选标签）
        """,
    )
    parser.add_argument("--product", type=str, default=None,
                        help="产品名（对应 products/<产品名>.json）；省略则用第一个可用产品")
    parser.add_argument("--tag", type=str, help="指定标签名（逗号分隔多个）")
    parser.add_argument("--all", action="store_true", help="采集全部37个标签")
    parser.add_argument("--rec", action="store_true", help="采集当前产品的推荐标签")
    parser.add_argument("--dry-run", action="store_true", help="预览模式：仅显示将要采集的标签，不实际执行")

    args = parser.parse_args()

    product = load_product(args.product)

    tags = resolve_tags(args, product)
    if not tags:
        print("未选择任何标签，退出。")
        return

    print(f"\n🎯 产品「{product.name}」将采集 {len(tags)} 个标签：")
    rec = set(product.recommended_tags or [])
    for i, tag in enumerate(tags, 1):
        marker = "⭐" if tag in rec else "  "
        print(f"  {marker} {i}. {tag}")

    if args.dry_run:
        print("\n[预览模式] 不会实际执行采集。")
        return

    print()

    scraper = YanajiaoScraper(product.name)
    try:
        scraper.run(tags)
    except KeyboardInterrupt:
        log("\n⚠️ 用户中断 (Ctrl+C)")
        scraper.close()
    except Exception as e:
        log(f"\n❌ 致命错误: {e}")
        import traceback
        traceback.print_exc()
        scraper.close()


if __name__ == "__main__":
    main()
